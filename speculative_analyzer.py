#!/usr/bin/env python3
"""
===============================================================================
SPECULATIVE ANALYZER - Trading de Corto Plazo (Long & Short)
===============================================================================
Script para análisis especulativo de corto plazo basado en:
    1. Análisis técnico (RSI, MACD, Bollinger, Ichimoku, Fibonacci)
    2. Noticias y sentimiento
    3. Señales de Polymarket (smart money)
    4. Trades de congresistas

GENERA HOJAS ADICIONALES EN EL EXCEL PRINCIPAL:
    - Technical_Signals: Indicadores técnicos por ticker
    - News_Sentiment: Noticias con análisis de sentimiento
    - Short_Term_Picks: Candidatos para trades corto plazo (LONG/SHORT)
    - Polymarket_Signals: Alertas de smart money
    - Congress_Trades: Trades recientes de congresistas

USO:
    python speculative_analyzer.py                    # Añade hojas al Excel existente
    python speculative_analyzer.py --excel custom.xlsx  # Especifica Excel
    python speculative_analyzer.py --standalone       # Genera Excel separado

DEPENDENCIAS:
    pip install pandas-ta finnhub-python vaderSentiment requests yfinance openpyxl

AUTOR: Claude (Anthropic)
FECHA: Enero 2026
===============================================================================
"""

import sys
import os
import argparse
from datetime import datetime
from typing import List, Dict, Optional

# Añadir path para imports locales
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Imports locales
from config.api_config import SHORT_TERM_WEIGHTS, TECHNICAL_CONFIG
from integrations.technical_indicators import TechnicalAnalyzer, analyze_batch
from integrations.news_analyzer import NewsAnalyzer, get_news_sentiment_batch
from integrations.polymarket_client import PolymarketClient
from integrations.congress_tracker import CongressTracker

# =============================================================================
# ESTILOS EXCEL
# =============================================================================
thin = Side(style='thin', color='CCCCCC')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

DARK_BLUE = '1F4E79'
DARK_GREEN = '006100'
DARK_RED = '9C0006'

header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
light_green_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
orange_fill = PatternFill(start_color='FBE4D5', end_color='FBE4D5', fill_type='solid')

white_font = Font(color='FFFFFF', bold=True, size=11)
title_font = Font(color=DARK_BLUE, bold=True, size=14)
green_font = Font(color=DARK_GREEN, bold=True)
red_font = Font(color=DARK_RED, bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =============================================================================
# UNIVERSO DE INVERSIÓN (hereda de market_analyzer o define propio)
# =============================================================================
def get_universe() -> List[str]:
    """Obtiene universo de tickers para análisis"""
    # Intentar cargar del Excel existente
    try:
        from market_analyzer import UNIVERSE
        tickers = []
        for sector_tickers in UNIVERSE.values():
            tickers.extend(sector_tickers)
        return tickers
    except ImportError:
        pass

    # Universo por defecto
    return [
        # Semiconductores
        'NVDA', 'TSM', 'AMD', 'ASML', 'AVGO', 'QCOM',
        # IA/Tech
        'ORCL', 'CRM', 'NOW', 'PLTR',
        # Defensa
        'LMT', 'NOC', 'RTX',
        # Energía
        'CVX', 'XOM', 'COP',
        # Defensivos
        'JNJ', 'PG', 'KO',
        # Pharma
        'NVO', 'MRK', 'LLY',
        # Streaming
        'NFLX', 'DIS',
        # China
        'BABA', 'PDD',
    ]


# =============================================================================
# ANÁLISIS PRINCIPAL
# =============================================================================
class SpeculativeAnalyzer:
    """
    Analizador especulativo para trading de corto plazo.
    """

    def __init__(self, tickers: List[str] = None):
        self.tickers = tickers or get_universe()
        self.tech_analyzer = TechnicalAnalyzer()
        self.news_analyzer = NewsAnalyzer()
        self.polymarket_client = PolymarketClient()
        self.congress_tracker = CongressTracker()

        # Resultados
        self.tech_results = []
        self.news_results = []
        self.polymarket_alerts = []
        self.congress_trades = []
        self.short_term_picks = []

    def analyze_technical(self, period: str = '6mo') -> List[Dict]:
        """Análisis técnico de todos los tickers"""
        print("\n📊 ANÁLISIS TÉCNICO")
        print("=" * 50)

        results = []
        total = len(self.tickers)

        for i, ticker in enumerate(self.tickers, 1):
            print(f"  [{i}/{total}] {ticker}...", end=" ")

            try:
                result = self.tech_analyzer.analyze(ticker, period)
                if result:
                    results.append(result)
                    direction = result.get('direction', 'WAIT')
                    score = result.get('tech_score', 50)
                    print(f"OK (Score: {score:.1f}, {direction})")
                else:
                    print("SKIP")
            except Exception as e:
                print(f"ERROR: {e}")

        self.tech_results = results
        print(f"\n✅ Análisis técnico completado: {len(results)}/{total} tickers")
        return results

    def analyze_news(self, days: int = 7) -> List[Dict]:
        """Análisis de noticias y sentimiento"""
        print("\n📰 ANÁLISIS DE NOTICIAS")
        print("=" * 50)

        results = []
        total = len(self.tickers)

        for i, ticker in enumerate(self.tickers, 1):
            print(f"  [{i}/{total}] {ticker}...", end=" ")

            try:
                sentiment = self.news_analyzer.get_sentiment_score(ticker, days)
                if sentiment:
                    results.append(sentiment)
                    signal = sentiment.get('signal', 'NO_DATA')
                    score = sentiment.get('score', 50)
                    print(f"OK (Score: {score:.1f}, {signal})")
                else:
                    print("SKIP")
            except Exception as e:
                print(f"ERROR: {e}")

        self.news_results = results
        print(f"\n✅ Análisis de noticias completado: {len(results)}/{total} tickers")
        return results

    def analyze_polymarket(self) -> List[Dict]:
        """Detecta alertas de smart money en Polymarket"""
        print("\n🎰 ANÁLISIS POLYMARKET (Smart Money)")
        print("=" * 50)

        try:
            alerts = self.polymarket_client.detect_smart_money_alerts()
            self.polymarket_alerts = alerts
            print(f"✅ Encontradas {len(alerts)} alertas de smart money")

            for alert in alerts[:5]:
                level = alert.get('alert_level', 'LOW')
                market = alert.get('market', '')[:50]
                volume = alert.get('volume_24h', '$0')
                print(f"  [{level}] {market}... ({volume})")

        except Exception as e:
            print(f"  ERROR: {e}")
            self.polymarket_alerts = []

        return self.polymarket_alerts

    def analyze_congress(self, days: int = 30) -> List[Dict]:
        """Obtiene trades recientes de congresistas"""
        print("\n🏛️ ANÁLISIS CONGRESS TRADES")
        print("=" * 50)

        try:
            # Trades de high performers
            trades = self.congress_tracker.get_high_performer_trades(days)
            self.congress_trades = self.congress_tracker.generate_excel_data(days)

            print(f"✅ Encontrados {len(self.congress_trades)} trades en últimos {days} días")

            # Top tickers
            top_tickers = self.congress_tracker.get_top_traded_tickers(days, limit=5)
            if not top_tickers.empty:
                print("\n  Top tickers operados por congresistas:")
                for _, row in top_tickers.iterrows():
                    print(f"    {row['ticker']}: {row['total_trades']} trades")

        except Exception as e:
            print(f"  ERROR: {e}")
            self.congress_trades = []

        return self.congress_trades

    def generate_short_term_picks(self) -> List[Dict]:
        """
        Genera picks de corto plazo combinando todos los análisis.
        """
        print("\n🎯 GENERANDO SHORT-TERM PICKS")
        print("=" * 50)

        picks = []
        weights = SHORT_TERM_WEIGHTS

        # Crear dict de resultados por ticker
        tech_by_ticker = {r['ticker']: r for r in self.tech_results}
        news_by_ticker = {r['ticker']: r for r in self.news_results}

        # Obtener señales de congress por ticker
        congress_signals = {}
        for ticker in self.tickers:
            try:
                signal = self.congress_tracker.get_signal_for_ticker(ticker, days=30)
                congress_signals[ticker] = signal
            except Exception:
                congress_signals[ticker] = {'score': 50, 'signal': 'NEUTRAL'}

        # Obtener señales de polymarket por ticker
        polymarket_signals = {}
        for ticker in self.tickers:
            try:
                signal = self.polymarket_client.get_signal_for_ticker(ticker)
                polymarket_signals[ticker] = signal
            except Exception:
                polymarket_signals[ticker] = {'score': 50, 'signal': 'NEUTRAL'}

        for ticker in self.tickers:
            tech = tech_by_ticker.get(ticker, {})
            news = news_by_ticker.get(ticker, {})
            congress = congress_signals.get(ticker, {})
            polymarket = polymarket_signals.get(ticker, {})

            if not tech:
                continue

            # Scores individuales
            tech_score = tech.get('tech_score', 50)
            news_score = news.get('score', 50)
            congress_score = congress.get('score', 50)
            polymarket_score = polymarket.get('score', 50)

            # Momentum (usar tech momentum)
            momentum_score = tech_score  # Simplificación

            # Score combinado
            combined_score = (
                tech_score * weights.get('technical', 0.35) +
                momentum_score * weights.get('momentum', 0.25) +
                news_score * weights.get('news_sentiment', 0.20) +
                congress_score * weights.get('congress', 0.10) +
                polymarket_score * weights.get('polymarket', 0.10)
            )

            # Determinar dirección
            direction = tech.get('direction', 'WAIT')
            if combined_score >= 60 and direction == 'LONG':
                direction = 'LONG'
            elif combined_score <= 40 and direction == 'SHORT':
                direction = 'SHORT'
            elif combined_score >= 55:
                direction = 'LONG'
            elif combined_score <= 45:
                direction = 'SHORT'
            else:
                direction = 'WAIT'

            # Confianza
            confidences = [
                tech.get('confidence', 'LOW'),
                news.get('confidence', 'LOW'),
                congress.get('confidence', 'LOW'),
            ]
            high_count = confidences.count('HIGH')
            medium_count = confidences.count('MEDIUM')

            if high_count >= 2:
                confidence = 'HIGH'
            elif high_count >= 1 or medium_count >= 2:
                confidence = 'MEDIUM'
            else:
                confidence = 'LOW'

            # Catalyst
            catalysts = []
            if tech.get('rsi_signal') in ['OVERSOLD', 'OVERBOUGHT']:
                catalysts.append(f"RSI {tech.get('rsi_signal')}")
            if tech.get('macd_signal_type') in ['BULLISH_CROSS', 'BEARISH_CROSS']:
                catalysts.append(f"MACD {tech.get('macd_signal_type')}")
            if news.get('high_impact_news'):
                catalysts.append("High-impact news")
            if congress.get('high_performer_activity'):
                catalysts.append("Congress activity")
            if polymarket.get('relevant_markets'):
                catalysts.append("Polymarket signal")

            picks.append({
                'ticker': ticker,
                'direction': direction,
                'combined_score': round(combined_score, 1),
                'tech_score': round(tech_score, 1),
                'news_score': round(news_score, 1),
                'congress_score': round(congress_score, 1),
                'polymarket_score': round(polymarket_score, 1),
                'price': tech.get('price', 0),
                'entry': tech.get('entry', tech.get('price', 0)),
                'stop_loss': tech.get('stop_loss', 0),
                'target_1': tech.get('target_1', 0),
                'target_2': tech.get('target_2', 0),
                'risk_reward': tech.get('risk_reward', 0),
                'confidence': confidence,
                'catalyst': ' | '.join(catalysts[:3]) if catalysts else 'Technical setup',
                'rsi': tech.get('rsi_14', 50),
                'ichimoku': tech.get('ichimoku_signal', 'N/A'),
            })

        # Ordenar por score y dirección
        picks.sort(key=lambda x: (
            0 if x['direction'] == 'WAIT' else 1,
            -x['combined_score'] if x['direction'] == 'LONG' else x['combined_score']
        ), reverse=True)

        self.short_term_picks = picks

        # Resumen
        longs = [p for p in picks if p['direction'] == 'LONG']
        shorts = [p for p in picks if p['direction'] == 'SHORT']
        waits = [p for p in picks if p['direction'] == 'WAIT']

        print(f"\n📊 Resumen:")
        print(f"  🟢 LONG: {len(longs)}")
        print(f"  🔴 SHORT: {len(shorts)}")
        print(f"  ⚪ WAIT: {len(waits)}")

        if longs:
            print(f"\n🟢 Top LONG picks:")
            for p in longs[:5]:
                print(f"    {p['ticker']:6} | Score: {p['combined_score']:5.1f} | {p['catalyst'][:30]}")

        if shorts:
            print(f"\n🔴 Top SHORT picks:")
            for p in shorts[:5]:
                print(f"    {p['ticker']:6} | Score: {p['combined_score']:5.1f} | {p['catalyst'][:30]}")

        return picks

    def run_full_analysis(self) -> Dict:
        """Ejecuta análisis completo"""
        print("\n" + "=" * 60)
        print("🚀 SPECULATIVE ANALYZER - ANÁLISIS COMPLETO")
        print("=" * 60)
        print(f"⏰ Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📈 Tickers: {len(self.tickers)}")

        # Ejecutar análisis
        self.analyze_technical()
        self.analyze_news()
        self.analyze_polymarket()
        self.analyze_congress()
        self.generate_short_term_picks()

        return {
            'tech_results': self.tech_results,
            'news_results': self.news_results,
            'polymarket_alerts': self.polymarket_alerts,
            'congress_trades': self.congress_trades,
            'short_term_picks': self.short_term_picks,
        }


# =============================================================================
# GENERACIÓN DE EXCEL
# =============================================================================
def create_technical_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Crea hoja Technical_Signals"""
    if 'Technical_Signals' in wb.sheetnames:
        del wb['Technical_Signals']

    ws = wb.create_sheet('Technical_Signals')
    ws['A1'] = f'📊 TECHNICAL SIGNALS - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:N1')

    headers = ['Ticker', 'Price', 'RSI', 'RSI_Signal', 'MACD_Cross', 'BB%',
               'Ichimoku', 'TK_Cross', 'Tech_Score', 'Direction', 'Signal',
               'Entry', 'Stop_Loss', 'Target_1', 'R:R']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center

    for i, r in enumerate(results, 1):
        row = i + 3

        ws.cell(row=row, column=1, value=r.get('ticker', '')).font = Font(bold=True)
        ws.cell(row=row, column=2, value=r.get('price', 0))
        ws.cell(row=row, column=3, value=r.get('rsi_14', ''))
        ws.cell(row=row, column=4, value=r.get('rsi_signal', ''))
        ws.cell(row=row, column=5, value=r.get('macd_signal_type', ''))
        ws.cell(row=row, column=6, value=r.get('bb_percent', ''))
        ws.cell(row=row, column=7, value=r.get('ichimoku_signal', ''))
        ws.cell(row=row, column=8, value=r.get('tk_cross', ''))

        # Tech score con color
        cell_score = ws.cell(row=row, column=9, value=r.get('tech_score', 50))
        score = r.get('tech_score', 50)
        if score >= 60:
            cell_score.fill = green_fill
        elif score <= 40:
            cell_score.fill = red_fill

        # Direction con color
        direction = r.get('direction', 'WAIT')
        cell_dir = ws.cell(row=row, column=10, value=direction)
        if direction == 'LONG':
            cell_dir.fill = green_fill
            cell_dir.font = green_font
        elif direction == 'SHORT':
            cell_dir.fill = red_fill
            cell_dir.font = red_font

        ws.cell(row=row, column=11, value=r.get('overall_signal', ''))
        ws.cell(row=row, column=12, value=r.get('entry', ''))
        ws.cell(row=row, column=13, value=r.get('stop_loss', ''))
        ws.cell(row=row, column=14, value=r.get('target_1', ''))
        ws.cell(row=row, column=15, value=r.get('risk_reward', ''))

        for col in range(1, 16):
            ws.cell(row=row, column=col).border = thin_border

    set_col_widths(ws, [8, 10, 6, 12, 14, 6, 14, 10, 10, 10, 12, 10, 10, 10, 6])


def create_news_sheet(wb: Workbook, analyzer: NewsAnalyzer, tickers: List[str]) -> None:
    """Crea hoja News_Sentiment"""
    if 'News_Sentiment' in wb.sheetnames:
        del wb['News_Sentiment']

    ws = wb.create_sheet('News_Sentiment')
    ws['A1'] = f'📰 NEWS SENTIMENT - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')

    headers = ['Ticker', 'Date', 'Headline', 'Source', 'Sentiment', 'Label',
               'Category', 'Impact', 'URL']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font

    row = 4
    for ticker in tickers[:20]:  # Limitar para no sobrecargar
        news_data = analyzer.generate_excel_data(ticker, days=7)
        for n in news_data[:5]:  # Max 5 noticias por ticker
            ws.cell(row=row, column=1, value=n.get('Ticker', ''))
            ws.cell(row=row, column=2, value=n.get('Date', ''))
            ws.cell(row=row, column=3, value=n.get('Headline', '')[:80])
            ws.cell(row=row, column=4, value=n.get('Source', ''))

            # Sentiment con color
            sentiment = n.get('Sentiment_Score', 0)
            cell_sent = ws.cell(row=row, column=5, value=sentiment)
            if sentiment > 0.1:
                cell_sent.fill = green_fill
            elif sentiment < -0.1:
                cell_sent.fill = red_fill

            ws.cell(row=row, column=6, value=n.get('Sentiment_Label', ''))
            ws.cell(row=row, column=7, value=n.get('Category', ''))

            # Impact con color
            impact = n.get('Impact', 'LOW')
            cell_impact = ws.cell(row=row, column=8, value=impact)
            if impact == 'HIGH':
                cell_impact.fill = orange_fill

            ws.cell(row=row, column=9, value=n.get('URL', ''))

            for col in range(1, 10):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    set_col_widths(ws, [8, 16, 60, 15, 10, 14, 12, 8, 50])


def create_picks_sheet(wb: Workbook, picks: List[Dict]) -> None:
    """Crea hoja Short_Term_Picks"""
    if 'Short_Term_Picks' in wb.sheetnames:
        del wb['Short_Term_Picks']

    ws = wb.create_sheet('Short_Term_Picks')
    ws['A1'] = f'🎯 SHORT-TERM PICKS - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:L1')

    # LONG section
    ws['A3'] = '🟢 LONG OPPORTUNITIES'
    ws['A3'].font = Font(bold=True, size=12, color=DARK_GREEN)
    ws['A3'].fill = green_fill

    headers = ['#', 'Ticker', 'Direction', 'Score', 'Tech', 'News', 'Congress',
               'Entry', 'Stop', 'Target', 'R:R', 'Catalyst']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = light_green_fill
        cell.font = Font(bold=True)

    row = 5
    longs = [p for p in picks if p['direction'] == 'LONG']
    for i, p in enumerate(longs[:15], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p['ticker']).font = Font(bold=True)
        ws.cell(row=row, column=3, value=p['direction'])
        ws.cell(row=row, column=4, value=p['combined_score'])
        ws.cell(row=row, column=5, value=p['tech_score'])
        ws.cell(row=row, column=6, value=p['news_score'])
        ws.cell(row=row, column=7, value=p['congress_score'])
        ws.cell(row=row, column=8, value=f"${p['entry']:.2f}" if p['entry'] else '')
        ws.cell(row=row, column=9, value=f"${p['stop_loss']:.2f}" if p['stop_loss'] else '')
        ws.cell(row=row, column=10, value=f"${p['target_1']:.2f}" if p['target_1'] else '')
        ws.cell(row=row, column=11, value=p['risk_reward'])
        ws.cell(row=row, column=12, value=p['catalyst'][:40])

        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # SHORT section
    row += 2
    ws.cell(row=row, column=1, value='🔴 SHORT OPPORTUNITIES')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=DARK_RED)
    ws.cell(row=row, column=1).fill = red_fill
    row += 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = red_fill
        cell.font = Font(bold=True, color='FFFFFF')
    row += 1

    shorts = [p for p in picks if p['direction'] == 'SHORT']
    for i, p in enumerate(shorts[:15], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p['ticker']).font = Font(bold=True)
        ws.cell(row=row, column=3, value=p['direction'])
        ws.cell(row=row, column=4, value=p['combined_score'])
        ws.cell(row=row, column=5, value=p['tech_score'])
        ws.cell(row=row, column=6, value=p['news_score'])
        ws.cell(row=row, column=7, value=p['congress_score'])
        ws.cell(row=row, column=8, value=f"${p['entry']:.2f}" if p['entry'] else '')
        ws.cell(row=row, column=9, value=f"${p['stop_loss']:.2f}" if p['stop_loss'] else '')
        ws.cell(row=row, column=10, value=f"${p['target_1']:.2f}" if p['target_1'] else '')
        ws.cell(row=row, column=11, value=p['risk_reward'])
        ws.cell(row=row, column=12, value=p['catalyst'][:40])

        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    set_col_widths(ws, [4, 8, 10, 8, 8, 8, 10, 10, 10, 10, 6, 40])


def create_polymarket_sheet(wb: Workbook, alerts: List[Dict]) -> None:
    """Crea hoja Polymarket_Signals"""
    if 'Polymarket_Signals' in wb.sheetnames:
        del wb['Polymarket_Signals']

    ws = wb.create_sheet('Polymarket_Signals')
    ws['A1'] = f'🎰 POLYMARKET SMART MONEY ALERTS - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')

    headers = ['Alert', 'Market', 'Volume_24h', 'Tickers', 'Impact', 'Action', 'Timestamp']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font

    for i, alert in enumerate(alerts, 1):
        row = i + 3

        # Alert level con color
        level = alert.get('alert_level', 'LOW')
        cell_level = ws.cell(row=row, column=1, value=level)
        if level == 'HIGH':
            cell_level.fill = red_fill
            cell_level.font = red_font
        elif level == 'MEDIUM':
            cell_level.fill = orange_fill

        ws.cell(row=row, column=2, value=alert.get('market', '')[:60])
        ws.cell(row=row, column=3, value=alert.get('volume_24h', ''))
        ws.cell(row=row, column=4, value=', '.join(alert.get('relevant_tickers', [])))
        ws.cell(row=row, column=5, value=alert.get('impact_assessment', '')[:40])
        ws.cell(row=row, column=6, value=alert.get('action_suggested', '')[:50])
        ws.cell(row=row, column=7, value=alert.get('timestamp', '')[:16])

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border

    set_col_widths(ws, [8, 50, 12, 20, 35, 45, 18])


def create_congress_sheet(wb: Workbook, trades: List[Dict]) -> None:
    """Crea hoja Congress_Trades"""
    if 'Congress_Trades' in wb.sheetnames:
        del wb['Congress_Trades']

    ws = wb.create_sheet('Congress_Trades')
    ws['A1'] = f'🏛️ CONGRESS TRADES - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:K1')

    headers = ['Politician', 'Party', 'Chamber', 'Ticker', 'Date', 'Type',
               'Amount', 'Asset', 'High_Perf', 'Signal']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font

    for i, trade in enumerate(trades[:100], 1):  # Limitar a 100
        row = i + 3

        ws.cell(row=row, column=1, value=trade.get('Politician', ''))
        ws.cell(row=row, column=2, value=trade.get('Party', ''))
        ws.cell(row=row, column=3, value=trade.get('Chamber', ''))
        ws.cell(row=row, column=4, value=trade.get('Ticker', '')).font = Font(bold=True)
        ws.cell(row=row, column=5, value=trade.get('Transaction_Date', ''))

        # Type con color
        tx_type = trade.get('Type', '')
        cell_type = ws.cell(row=row, column=6, value=tx_type)
        if tx_type == 'PURCHASE':
            cell_type.fill = green_fill
        elif tx_type == 'SALE':
            cell_type.fill = red_fill

        ws.cell(row=row, column=7, value=trade.get('Amount_Range', ''))
        ws.cell(row=row, column=8, value=trade.get('Asset_Type', ''))

        # High performer con color
        is_hp = trade.get('Is_High_Performer', 'No')
        cell_hp = ws.cell(row=row, column=9, value=is_hp)
        if is_hp == 'Yes':
            cell_hp.fill = yellow_fill
            cell_hp.font = Font(bold=True)

        ws.cell(row=row, column=10, value=trade.get('Signal_Strength', ''))

        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    set_col_widths(ws, [20, 6, 8, 8, 12, 10, 18, 8, 10, 8])


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description='Speculative Analyzer - Short-Term Trading')
    parser.add_argument('--excel', default='market_analysis_auto.xlsx',
                        help='Excel file to add sheets to')
    parser.add_argument('--standalone', action='store_true',
                        help='Create standalone Excel instead of adding to existing')
    parser.add_argument('--tickers', nargs='+',
                        help='Specific tickers to analyze')

    args = parser.parse_args()

    # Tickers
    tickers = args.tickers if args.tickers else get_universe()

    print("\n" + "=" * 60)
    print("🚀 SPECULATIVE ANALYZER v1.0")
    print("=" * 60)

    # Ejecutar análisis
    analyzer = SpeculativeAnalyzer(tickers)
    results = analyzer.run_full_analysis()

    # Crear/abrir Excel
    print("\n📝 GENERANDO EXCEL")
    print("=" * 50)

    if args.standalone:
        wb = Workbook()
        output_path = 'speculative_analysis.xlsx'
    else:
        try:
            wb = load_workbook(args.excel)
            output_path = args.excel
            print(f"  Añadiendo hojas a {args.excel}")
        except FileNotFoundError:
            wb = Workbook()
            output_path = args.excel
            print(f"  Creando nuevo archivo {args.excel}")

    # Crear hojas
    print("  Creando Technical_Signals...")
    create_technical_sheet(wb, results['tech_results'])

    print("  Creando News_Sentiment...")
    create_news_sheet(wb, analyzer.news_analyzer, tickers)

    print("  Creando Short_Term_Picks...")
    create_picks_sheet(wb, results['short_term_picks'])

    print("  Creando Polymarket_Signals...")
    create_polymarket_sheet(wb, results['polymarket_alerts'])

    print("  Creando Congress_Trades...")
    create_congress_sheet(wb, results['congress_trades'])

    # Guardar
    wb.save(output_path)
    print(f"\n✅ Excel guardado: {output_path}")

    # Resumen final
    print("\n" + "=" * 60)
    print("✅ ANÁLISIS COMPLETADO")
    print("=" * 60)

    longs = [p for p in results['short_term_picks'] if p['direction'] == 'LONG']
    shorts = [p for p in results['short_term_picks'] if p['direction'] == 'SHORT']

    print(f"\n📊 Resumen:")
    print(f"  🟢 LONG opportunities: {len(longs)}")
    print(f"  🔴 SHORT opportunities: {len(shorts)}")
    print(f"  🎰 Polymarket alerts: {len(results['polymarket_alerts'])}")
    print(f"  🏛️ Congress trades: {len(results['congress_trades'])}")

    if longs:
        print(f"\n🎯 Top 3 LONG:")
        for p in longs[:3]:
            print(f"    {p['ticker']:6} | Score: {p['combined_score']:5.1f} | Entry: ${p['entry']:.2f}")

    if shorts:
        print(f"\n🎯 Top 3 SHORT:")
        for p in shorts[:3]:
            print(f"    {p['ticker']:6} | Score: {p['combined_score']:5.1f} | Entry: ${p['entry']:.2f}")

    print(f"\n📁 Revisa las hojas en: {output_path}")
    print("   - Technical_Signals: Indicadores técnicos")
    print("   - News_Sentiment: Noticias y sentimiento")
    print("   - Short_Term_Picks: Picks de corto plazo")
    print("   - Polymarket_Signals: Alertas smart money")
    print("   - Congress_Trades: Trades de congresistas")


if __name__ == '__main__':
    main()
