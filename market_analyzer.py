#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
MARKET ANALYZER V13 - SISTEMA SEMI-AUTOMATIZADO
═══════════════════════════════════════════════════════════════════════════════
Script de actualización automática usando yfinance (GRATIS, sin API key)

USO:
    python market_analyzer.py                    # Genera Excel con datos actualizados
    python market_analyzer.py --regime CONTRACTION  # Cambia régimen macro
    python market_analyzer.py --add AAPL GOOGL   # Añade tickers al universo

REQUISITOS:
    pip install yfinance openpyxl pandas numpy

DATOS QUE OBTIENE YFINANCE (GRATIS):
    ✅ Precio actual
    ✅ Forward P/E
    ✅ Trailing P/E
    ✅ EV/EBITDA
    ✅ Beta
    ✅ 52-week high/low
    ✅ Target price (consenso analistas)
    ✅ Analyst recommendations
    ✅ Profit margins
    ✅ ROE
    ✅ Dividend yield
    ✅ Historical prices (para calcular momentum y volatilidad)

LIMITACIONES:
    ⚠️ ROIC no disponible directamente (se aproxima con ROE)
    ⚠️ FCF Yield requiere cálculo manual
    ⚠️ Analyst revisions no disponible (se usa recommendation changes)
    ⚠️ Algunos datos pueden estar desactualizados 1-2 días

AUTOR: Claude (Anthropic) para Peter
FECHA: Enero 2026

V13 UPDATES:
    - Integración con Polymarket (smart money detection)
    - Integración con Congress Tracker (trades de congresistas)
    - Sistema de scoring mejorado con 6 factores
═══════════════════════════════════════════════════════════════════════════════
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from urllib.parse import quote
from datetime import datetime, timedelta
from collections import Counter
import argparse
import sys
import time
import io

# Fix encoding para Windows (emojis) + line buffering para ver output en tiempo real
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)

# Imports de integraciones (opcionales - funcionan sin ellas)
try:
    from integrations.congress_tracker import CongressTracker
    CONGRESS_AVAILABLE = True
except ImportError:
    CONGRESS_AVAILABLE = False

try:
    from integrations.polymarket_client import PolymarketClient
    POLYMARKET_AVAILABLE = True
except ImportError:
    POLYMARKET_AVAILABLE = False

try:
    from integrations.monetary_plumbing import (
        analyze_monetary_regime,
        get_long_term_weights,
        get_momentum_weights,
        get_optimal_weights_for_regime
    )
    MONETARY_AVAILABLE = True
except ImportError:
    MONETARY_AVAILABLE = False

try:
    from integrations.trigger_system import (
        TriggerSystem,
        run_full_scan,
        STRATEGY_RECOMMENDATIONS,
        CORRELATIONS
    )
    TRIGGER_SYSTEM_AVAILABLE = True
except ImportError:
    TRIGGER_SYSTEM_AVAILABLE = False

try:
    from integrations.signal_database import SignalDatabase
    SIGNAL_DB_AVAILABLE = True
except ImportError:
    SIGNAL_DB_AVAILABLE = False

try:
    from integrations.investment_thesis import (
        InvestmentThesisGenerator,
        generate_thesis_batch,
        STRATEGY_BY_HORIZON,
        DATA_SOURCES
    )
    THESIS_GENERATOR_AVAILABLE = True
except ImportError:
    THESIS_GENERATOR_AVAILABLE = False

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

# Universo de inversión - EDITAR AQUÍ PARA AÑADIR/QUITAR TICKERS
# Expandido con empresas internacionales de Europa, Asia, Latam
UNIVERSE = {
    # ═══════════════════════════════════════════════════════════════
    # TECNOLOGÍA
    # ═══════════════════════════════════════════════════════════════
    'Semiconductores': [
        'NVDA', 'TSM', 'AMD', 'ASML', 'AVGO', 'QCOM', 'INTC', 'MU',  # USA/TW/NL
        'SSNLF',  # Samsung (Korea) - OTC
        'IFNNY',  # Infineon (Germany) - ADR
        'TOELY',  # Tokyo Electron (Japan) - ADR
    ],
    'IA Software': [
        'ORCL', 'CRM', 'SAP', 'IBM', 'NOW', 'PLTR',  # USA/Germany
        'DDOG', 'SNOW', 'MDB',  # Cloud/Data USA
        'PATH', 'AI',  # Automation USA
    ],
    'Big Tech': [
        'AAPL', 'MSFT', 'GOOGL', 'META', 'AMZN',  # USA Mega caps
    ],
    'Fintech': [
        'V', 'MA', 'PYPL', 'XYZ',  # USA
        'NU',  # Nubank (Brazil/Latam)
        'MELI',  # MercadoLibre (Latam)
        'SE',  # Sea Limited (Singapore/SEA)
        'AFRM',  # Affirm (BNPL)
    ],

    # ═══════════════════════════════════════════════════════════════
    # ASIA
    # ═══════════════════════════════════════════════════════════════
    'China Tech': [
        'PDD', 'BABA', 'JD', 'BIDU',  # E-commerce/Search
        'TCEHY',  # Tencent ADR
        'NTES',  # NetEase Gaming
        'LI', 'NIO', 'XPEV',  # EVs China
    ],
    'Japan': [
        'TM',  # Toyota
        'SONY',  # Sony
        'MUFG',  # Mitsubishi UFJ Financial
        'NTDOY',  # Nintendo
        'SMFG',  # Sumitomo Mitsui
        'HMC',  # Honda
    ],
    'Korea': [
        'SSNLF',  # Samsung OTC
        'LPL',  # LG Display
        'KB',  # KB Financial
        'PKX',  # POSCO Steel
    ],
    'India': [
        'IBN',  # ICICI Bank
        'INFY',  # Infosys
        'WIT',  # Wipro
        'HDB',  # HDFC Bank
        'RDY',  # Dr. Reddy's Labs ADR
    ],
    'SEA': [
        'SE',  # Sea Limited (Singapore)
        'GRAB',  # Grab Holdings (Singapore)
    ],

    # ═══════════════════════════════════════════════════════════════
    # EUROPA
    # ═══════════════════════════════════════════════════════════════
    'Lujo': [
        'MC.PA',  # LVMH (France)
        'RACE',  # Ferrari (Italy)
        'RMS.PA',  # Hermès (France)
        'KER.PA',  # Kering (France)
        'BRBY.L',  # Burberry (UK)
    ],
    'Europa Industriales': [
        'SIE.DE',  # Siemens (Germany)
        'AIR.PA',  # Airbus (France)
        'ABBNY',  # ABB (Switzerland) ADR
        'ATLKY',  # Atlas Copco (Sweden) ADR
        'RELX',  # RELX (UK)
    ],
    'Europa Financieras': [
        'UBS',  # UBS (Switzerland) - absorbió Credit Suisse
        'HSBC',  # HSBC (UK)
        'BCS',  # Barclays (UK)
        'SAN',  # Santander (Spain)
        'BBVA',  # BBVA (Spain)
        'ING',  # ING (Netherlands)
        'DB',  # Deutsche Bank (Germany)
    ],
    'Europa Tech': [
        'ASML',  # ASML (Netherlands)
        'SAP',  # SAP (Germany)
        'SPOT',  # Spotify (Sweden)
        'NOK',  # Nokia (Finland) - ADR en NYSE
        'ERIC',  # Ericsson (Sweden)
    ],

    # ═══════════════════════════════════════════════════════════════
    # LATAM
    # ═══════════════════════════════════════════════════════════════
    'Latam': [
        'NU',  # Nubank (Brazil fintech)
        'MELI',  # MercadoLibre (Argentina/Latam)
        'VALE',  # Vale Mining (Brazil)
        'PBR',  # Petrobras (Brazil)
        'ITUB',  # Itaú Unibanco (Brazil)
        'BSBR',  # Banco Santander Brasil
        'AMX',  # América Móvil (Mexico)
        'CEMEX',  # Cemex (Mexico)
        'KOF',  # Coca-Cola Femsa (Mexico)
    ],

    # ═══════════════════════════════════════════════════════════════
    # DEFENSA Y AEROESPACIAL
    # ═══════════════════════════════════════════════════════════════
    'Defensa': [
        'LMT', 'NOC', 'GD', 'RTX', 'BA',  # USA
        'AIR.PA',  # Airbus (dual-use)
        'BA.L',  # BAE Systems (UK)
        'RNMBY',  # Rheinmetall (Germany) ADR
    ],

    # ═══════════════════════════════════════════════════════════════
    # ENERGÍA
    # ═══════════════════════════════════════════════════════════════
    'Energia Oil': [
        'CVX', 'XOM', 'COP',  # USA
        'TTE',  # TotalEnergies (France)
        'SHEL',  # Shell (UK/NL)
        'BP',  # BP (UK)
        'EQNR',  # Equinor (Norway)
        'PBR',  # Petrobras (Brazil)
        'SU',  # Suncor (Canada)
        'EC',  # Ecopetrol (Colombia)
    ],
    'Energia Renovables': [
        'ENPH', 'SEDG', 'FSLR',  # Solar USA
        'NEE', 'AES',  # Utilities USA
        'ORSTED.CO',  # Ørsted (Denmark)
        'IBDRY',  # Iberdrola (Spain) ADR
        'VWS.CO',  # Vestas Wind (Denmark)
    ],
    'Nuclear': [
        'CCJ',  # Cameco (Canada)
        'CEG',  # Constellation Energy (USA)
        'VST',  # Vistra (USA)
        'UEC',  # Uranium Energy Corp (USA)
        'UUUU',  # Energy Fuels (USA)
    ],

    # ═══════════════════════════════════════════════════════════════
    # MATERIALES
    # ═══════════════════════════════════════════════════════════════
    'Oro': [
        'NEM', 'GOLD', 'AEM', 'GFI',  # USA/Canada/SA
        'FNV', 'WPM',  # Royalty/Streaming
        'AU',  # AngloGold Ashanti
    ],
    'Cobre Litio': [
        'FCX', 'SCCO',  # USA/Peru
        'BHP', 'RIO',  # Australia/UK
        'VALE',  # Vale (Brazil)
        'ALB', 'LAC', 'SQM',  # Litio
        'TECK',  # Teck Resources (Canada)
    ],

    # ═══════════════════════════════════════════════════════════════
    # CONSUMO Y DEFENSIVO
    # ═══════════════════════════════════════════════════════════════
    'Defensivo': [
        'JNJ', 'PG', 'KO', 'PEP', 'WMT', 'COST',  # USA
        'NESN.SW',  # Nestlé (Switzerland)
        'UL',  # Unilever (UK/NL)
        'DEO',  # Diageo (UK)
    ],
    'Streaming Media': [
        'NFLX', 'DIS', 'SPOT',  # USA/Sweden
        'WBD',  # Warner Bros Discovery
        'CMCSA',  # Comcast (Peacock)
    ],

    # ═══════════════════════════════════════════════════════════════
    # SALUD
    # ═══════════════════════════════════════════════════════════════
    'Pharma': [
        'NVO',  # Novo Nordisk (Denmark)
        'MRK', 'LLY', 'PFE', 'ABBV', 'BMY',  # USA
        'AZN',  # AstraZeneca (UK)
        'GSK',  # GSK (UK)
        'SNY',  # Sanofi (France)
        'ROG.SW',  # Roche (Switzerland)
        'NOVN.SW',  # Novartis (Switzerland)
    ],
    'Biotech': [
        'VRTX', 'REGN', 'GILD', 'BIIB',  # USA Large
        'MRNA', 'BNTX',  # mRNA
    ],

    # ═══════════════════════════════════════════════════════════════
    # REAL ESTATE / REITS
    # ═══════════════════════════════════════════════════════════════
    'REITs': [
        'PLD', 'AMT', 'EQIX',  # Industrial/Data Centers USA
        'SPG', 'O',  # Retail USA
        'DLR',  # Digital Realty
    ],
}

# Mapeo de países por ticker (para ajustes regionales)
# Expandido para incluir todos los nuevos tickers internacionales
COUNTRY_MAP = {
    # Asia
    'TSM': ('Taiwan', 'Asia'),
    'SSNLF': ('Korea', 'Asia'),
    'TOELY': ('Japan', 'Asia'),
    'PDD': ('China', 'Asia'),
    'BABA': ('China', 'Asia'),
    'JD': ('China', 'Asia'),
    'BIDU': ('China', 'Asia'),
    'TCEHY': ('China', 'Asia'),
    'NTES': ('China', 'Asia'),
    'LI': ('China', 'Asia'),
    'NIO': ('China', 'Asia'),
    'XPEV': ('China', 'Asia'),
    'TM': ('Japan', 'Asia'),
    'SONY': ('Japan', 'Asia'),
    'MUFG': ('Japan', 'Asia'),
    'NTDOY': ('Japan', 'Asia'),
    'SMFG': ('Japan', 'Asia'),
    'HMC': ('Japan', 'Asia'),
    'LPL': ('Korea', 'Asia'),
    'KB': ('Korea', 'Asia'),
    'PKX': ('Korea', 'Asia'),
    'IBN': ('India', 'Asia'),
    'INFY': ('India', 'Asia'),
    'WIT': ('India', 'Asia'),
    'HDB': ('India', 'Asia'),
    'RDY': ('India', 'Asia'),
    'SE': ('Singapore', 'Asia'),
    'GRAB': ('Singapore', 'Asia'),
    'BHP': ('Australia', 'Oceania'),
    'RIO': ('Australia', 'Oceania'),

    # Europe
    'ASML': ('Netherlands', 'Europe'),
    'SAP': ('Germany', 'Europe'),
    'IFNNY': ('Germany', 'Europe'),
    'SIE.DE': ('Germany', 'Europe'),
    'RNMBY': ('Germany', 'Europe'),
    'MC.PA': ('France', 'Europe'),
    'RMS.PA': ('France', 'Europe'),
    'KER.PA': ('France', 'Europe'),
    'AIR.PA': ('France', 'Europe'),
    'TTE': ('France', 'Europe'),
    'SNY': ('France', 'Europe'),
    'RACE': ('Italy', 'Europe'),
    'SHEL': ('UK', 'Europe'),
    'BRBY.L': ('UK', 'Europe'),
    'BA.L': ('UK', 'Europe'),
    'RELX': ('UK', 'Europe'),
    'HSBC': ('UK', 'Europe'),
    'BCS': ('UK', 'Europe'),
    'BP': ('UK', 'Europe'),
    'AZN': ('UK', 'Europe'),
    'GSK': ('UK', 'Europe'),
    'DEO': ('UK', 'Europe'),
    'UL': ('UK', 'Europe'),
    'ABBNY': ('Switzerland', 'Europe'),
    'UBS': ('Switzerland', 'Europe'),
    'ROG.SW': ('Switzerland', 'Europe'),
    'NOVN.SW': ('Switzerland', 'Europe'),
    'NESN.SW': ('Switzerland', 'Europe'),
    'SPOT': ('Sweden', 'Europe'),
    'ATLKY': ('Sweden', 'Europe'),
    'ERIC': ('Sweden', 'Europe'),
    'SAN': ('Spain', 'Europe'),
    'BBVA': ('Spain', 'Europe'),
    'IBDRY': ('Spain', 'Europe'),
    'ING': ('Netherlands', 'Europe'),
    'NOK': ('Finland', 'Europe'),
    'DB': ('Germany', 'Europe'),
    'NVO': ('Denmark', 'Europe'),
    'ORSTED.CO': ('Denmark', 'Europe'),
    'VWS.CO': ('Denmark', 'Europe'),
    'EQNR': ('Norway', 'Europe'),
    'BNTX': ('Germany', 'Europe'),

    # Americas (non-USA)
    'GOLD': ('Canada', 'Americas'),
    'AEM': ('Canada', 'Americas'),
    'CCJ': ('Canada', 'Americas'),
    'FNV': ('Canada', 'Americas'),
    'WPM': ('Canada', 'Americas'),
    'SU': ('Canada', 'Americas'),
    'TECK': ('Canada', 'Americas'),
    'LAC': ('Canada', 'Americas'),
    'NU': ('Brazil', 'Latam'),
    'VALE': ('Brazil', 'Latam'),
    'PBR': ('Brazil', 'Latam'),
    'ITUB': ('Brazil', 'Latam'),
    'BSBR': ('Brazil', 'Latam'),
    'MELI': ('Argentina', 'Latam'),
    'AMX': ('Mexico', 'Latam'),
    'CEMEX': ('Mexico', 'Latam'),
    'KOF': ('Mexico', 'Latam'),
    'EC': ('Colombia', 'Latam'),
    'SQM': ('Chile', 'Latam'),
    'SCCO': ('Peru', 'Latam'),
    'AU': ('South Africa', 'Africa'),
    'GFI': ('South Africa', 'Africa'),
}

# Régimen macro - EDITAR SEGÚN CONDICIONES DE MERCADO
MACRO_CONFIG = {
    'regime': 'LATE_EXPANSION',  # RECOVERY, EXPANSION, LATE_EXPANSION, CONTRACTION
    'risk_appetite': 'NEUTRAL',   # RISK_ON, NEUTRAL, RISK_OFF
}

# Pesos base y ajustes
# V14: Pesos optimizados por backtesting (5 escenarios historicos)
# El backtesting mostro que Momentum Heavy funciona mejor en la mayoria de escenarios
BASE_WEIGHTS = {
    'value': 0.15,         # Reducido - menos importante en mercados de momentum
    'quality': 0.20,       # Reducido - pero sigue siendo clave para proteccion
    'momentum': 0.35,      # AUMENTADO - factor mas predictivo en backtest
    'lowvol': 0.10,        # Reducido - menos relevante excepto en crisis
    'congress': 0.10,      # Trades de congresistas (edge real)
    'polymarket': 0.10,    # Smart money de Polymarket
}

REGIME_ADJUSTMENTS = {
    'RECOVERY': {'value': +0.05, 'quality': -0.05, 'momentum': +0.05, 'lowvol': -0.05},
    'EXPANSION': {'value': +0.05, 'quality': 0, 'momentum': +0.05, 'lowvol': -0.10},
    'LATE_EXPANSION': {'value': -0.05, 'quality': +0.05, 'momentum': -0.05, 'lowvol': +0.05},
    'CONTRACTION': {'value': -0.05, 'quality': +0.10, 'momentum': -0.10, 'lowvol': +0.05},
}

RISK_ADJUSTMENTS = {
    'RISK_ON': {'value': 0, 'quality': -0.05, 'momentum': +0.10, 'lowvol': -0.05},
    'NEUTRAL': {'value': 0, 'quality': 0, 'momentum': 0, 'lowvol': 0},
    'RISK_OFF': {'value': 0, 'quality': +0.05, 'momentum': -0.10, 'lowvol': +0.05},
}

# =============================================================================
# ESTILOS EXCEL
# =============================================================================
thin = Side(style='thin', color='CCCCCC')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

DARK_BLUE = '1F4E79'
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
subheader_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
light_green_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
light_blue_fill = PatternFill(start_color='DBEEF7', end_color='DBEEF7', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

white_font = Font(color='FFFFFF', bold=True, size=11)
title_font = Font(color=DARK_BLUE, bold=True, size=16)
green_font = Font(color='006100', bold=True)
red_font = Font(color='9C0006', bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap = Alignment(wrap_text=True, vertical='top')

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# FUNCIONES DE DATOS
# =============================================================================

def get_final_weights(investment_type: str = 'balanced'):
    """
    Calcula pesos finales según régimen macro y análisis de liquidez.

    Args:
        investment_type: 'long_term', 'momentum', o 'balanced'
    """
    global MONETARY_ANALYSIS  # Para usar en otras partes del código

    # Analizar régimen monetario si está disponible
    if MONETARY_AVAILABLE:
        try:
            MONETARY_ANALYSIS = analyze_monetary_regime()

            if investment_type == 'long_term':
                return get_long_term_weights(MONETARY_ANALYSIS)
            elif investment_type == 'momentum':
                return get_momentum_weights(MONETARY_ANALYSIS)
            else:
                return get_optimal_weights_for_regime(BASE_WEIGHTS, MONETARY_ANALYSIS)
        except Exception as e:
            print(f"  Warning: Monetary analysis failed: {e}")
            MONETARY_ANALYSIS = None
    else:
        MONETARY_ANALYSIS = None

    # Fallback al método anterior
    weights = BASE_WEIGHTS.copy()
    regime_adj = REGIME_ADJUSTMENTS.get(MACRO_CONFIG['regime'], {})
    risk_adj = RISK_ADJUSTMENTS.get(MACRO_CONFIG['risk_appetite'], {})

    for factor in weights:
        weights[factor] += regime_adj.get(factor, 0) + risk_adj.get(factor, 0)

    total = sum(weights.values())
    return {k: v/total for k, v in weights.items()}

# Variable global para almacenar análisis monetario
MONETARY_ANALYSIS = None


def fetch_stock_data(ticker, sector):
    """
    Obtiene datos de un ticker usando yfinance
    Retorna diccionario con métricas o None si falla
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # Datos básicos
        price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
        if not price:
            print(f"  ⚠️ {ticker}: Sin precio disponible")
            return None
        
        name = info.get('shortName', ticker)
        
        # País y región
        country_info = COUNTRY_MAP.get(ticker, ('US', 'Americas'))
        country, region = country_info
        
        # Valoración
        fwd_pe = info.get('forwardPE')
        trailing_pe = info.get('trailingPE')
        ev_ebitda = info.get('enterpriseToEbitda')
        
        # Calidad
        roe = info.get('returnOnEquity', 0)
        if roe:
            roe = roe * 100  # Convertir a porcentaje
        
        op_margin = info.get('operatingMargins', 0)
        if op_margin:
            op_margin = op_margin * 100
        
        profit_margin = info.get('profitMargins', 0)
        if profit_margin:
            profit_margin = profit_margin * 100
        
        # Aproximar ROIC con ROE (simplificación)
        roic = roe * 0.8 if roe else None  # ROIC típicamente menor que ROE
        
        # Deuda
        total_debt = info.get('totalDebt', 0) or 0
        total_cash = info.get('totalCash', 0) or 0
        ebitda = info.get('ebitda', 1) or 1
        net_debt_ebitda = (total_debt - total_cash) / ebitda if ebitda > 0 else None
        
        # FCF Yield aproximado
        free_cf = info.get('freeCashflow', 0) or 0
        market_cap = info.get('marketCap', 1) or 1
        fcf_yield = (free_cf / market_cap) * 100 if market_cap > 0 else None
        
        # Riesgo
        beta = info.get('beta')
        
        # Dividendo
        div_yield = info.get('dividendYield', 0) or 0
        if div_yield:
            div_yield = div_yield * 100
        
        # Target y rating
        target = info.get('targetMeanPrice')
        
        # Recomendación (1=Strong Buy, 5=Sell)
        rec = info.get('recommendationMean', 3)
        # Convertir a escala 1-5 donde 5 es mejor
        rating = 6 - rec if rec else 3
        
        # Momentum - obtener histórico
        try:
            hist = stock.history(period='1y')
            if len(hist) > 0:
                price_1y_ago = hist['Close'].iloc[0]
                mom_12m = ((price - price_1y_ago) / price_1y_ago) * 100
                
                # Volatilidad
                returns = hist['Close'].pct_change().dropna()
                vol_252d = returns.std() * np.sqrt(252) * 100
            else:
                mom_12m = 0
                vol_252d = 30
        except:
            mom_12m = 0
            vol_252d = 30
        
        # Analyst revisions (aproximación con recommendation trend)
        # yfinance no tiene esto directamente, usamos cambio en target
        analyst_rev = 0  # Por defecto neutral
        
        return {
            'ticker': ticker,
            'name': name[:25],  # Truncar nombre largo
            'sector': sector,
            'country': country,
            'region': region,
            'price': round(price, 2) if price else None,
            'fwd_pe': round(fwd_pe, 1) if fwd_pe else None,
            'trailing_pe': round(trailing_pe, 1) if trailing_pe else None,
            'ev_ebitda': round(ev_ebitda, 1) if ev_ebitda else None,
            'fcf_yield': round(fcf_yield, 1) if fcf_yield else None,
            'roic': round(roic, 1) if roic else None,
            'roe': round(roe, 1) if roe else None,
            'op_margin': round(op_margin, 1) if op_margin else None,
            'net_debt_ebitda': round(net_debt_ebitda, 2) if net_debt_ebitda else None,
            'beta': round(beta, 2) if beta else 1.0,
            'vol_252d': round(vol_252d, 1) if vol_252d else 30,
            'mom_12m': round(mom_12m, 1) if mom_12m else 0,
            'analyst_rev': analyst_rev,
            'target': round(target, 2) if target else None,
            'rating': round(rating, 1) if rating else 3.0,
            'dividend_yield': round(div_yield, 2) if div_yield else 0,
        }
        
    except Exception as e:
        print(f"  ❌ {ticker}: Error - {str(e)[:50]}")
        return None


def fetch_all_data():
    """Obtiene datos de todo el universo (FIXED + TRIGGERED)"""
    global TRIGGERED_STOCKS

    companies = []
    total_tickers = sum(len(tickers) for tickers in UNIVERSE.values())

    print(f"\n  Descargando datos de {total_tickers} tickers...")
    print("  " + "-" * 50)

    count = 0
    for sector, tickers in UNIVERSE.items():
        is_triggered_sector = (sector == 'Triggered')
        print(f"\n  {sector} {'[TRIGGERED]' if is_triggered_sector else '[FIXED]'}:")

        for ticker in tickers:
            count += 1
            print(f"    [{count}/{total_tickers}] {ticker}...", end=" ")

            data = fetch_stock_data(ticker, sector)
            if data:
                # Marcar origen: FIXED (universo estático) o TRIGGERED (descubierto por triggers)
                if is_triggered_sector or ticker in TRIGGERED_STOCKS:
                    data['source'] = 'TRIGGERED'
                    # Obtener razón del trigger
                    triggers = TRIGGERED_STOCKS.get(ticker, [])
                    if triggers:
                        reasons = []
                        for t in triggers[:2]:  # Máx 2 razones
                            if t.get('type') == 'CORRELATION':
                                reasons.append(f"{t['source']} {t['direction']}")
                            elif t.get('type') == 'CONGRESS_TRADE':
                                reasons.append(f"{t['politician'][:12]} {t['action']}")
                            elif t.get('type') == 'MONETARY':
                                reasons.append(t['subtype'])
                        data['trigger_reason'] = ' | '.join(reasons) if reasons else 'Trigger activo'
                    else:
                        data['trigger_reason'] = 'En universo + trigger'
                else:
                    data['source'] = 'FIXED'
                    data['trigger_reason'] = 'Universo estático'

                companies.append(data)
                print(f"OK ${data['price']}")
            else:
                print("FAIL")

            # Pausa para evitar rate limiting
            time.sleep(0.3)

    print(f"\n  Datos obtenidos: {len(companies)}/{total_tickers} tickers")
    return companies


def percentile_rank(values, reverse=False):
    """Calcula percentiles (0-100)"""
    result = []
    valid_with_idx = [(i, v) for i, v in enumerate(values) if v is not None]
    if not valid_with_idx:
        return [50.0 for _ in values]
    
    sorted_vals = sorted(valid_with_idx, key=lambda x: x[1], reverse=reverse)
    ranks = {idx: ((rank + 1) / len(sorted_vals)) * 100 
             for rank, (idx, _) in enumerate(sorted_vals)}
    
    for i, v in enumerate(values):
        result.append(ranks.get(i, 50.0) if v is not None else 50.0)
    return result


def generate_investment_thesis(c, monetary_data=None, congress_details=None):
    """
    Genera una tesis de inversión profesional basada en los factores.
    Explica POR QUÉ se recomienda la acción tomada con DATOS CONCRETOS.

    Args:
        c: Diccionario con datos de la empresa
        monetary_data: Dict con régimen monetario (opcional)
        congress_details: Dict con detalles de trades de congresistas (opcional)
    """
    ticker = c.get('ticker', '')
    signal = c['signal']
    score = c['composite_score']
    upside = c.get('upside', 0)
    price = c.get('price', 0)
    target = c.get('target', 0)

    # Extraer datos concretos para la tesis
    fwd_pe = c.get('fwd_pe')
    ev_ebitda = c.get('ev_ebitda')
    roe = c.get('roe') or c.get('roic')
    op_margin = c.get('op_margin')
    mom_12m = c.get('mom_12m')
    beta = c.get('beta')
    vol = c.get('vol_252d')

    # Scores
    value_score = c.get('value_score', 50)
    quality_score = c.get('quality_score', 50)
    momentum_score = c.get('momentum_score', 50)
    lowvol_score = c.get('lowvol_score', 50)
    congress_score = c.get('congress_score', 50)
    poly_score = c.get('polymarket_score', 50)

    # Identificar fortalezas y debilidades CON DATOS
    strengths = []
    weaknesses = []

    # VALUE con datos
    if value_score >= 65:
        details = []
        if fwd_pe: details.append(f"P/E {fwd_pe:.1f}x")
        if ev_ebitda: details.append(f"EV/EBITDA {ev_ebitda:.1f}x")
        strengths.append(f"valoración atractiva ({', '.join(details) if details else 'percentil bajo'})")
    elif value_score <= 35:
        details = []
        if fwd_pe: details.append(f"P/E {fwd_pe:.1f}x")
        if ev_ebitda: details.append(f"EV/EBITDA {ev_ebitda:.1f}x")
        weaknesses.append(f"valoración cara ({', '.join(details) if details else 'percentil alto'})")

    # QUALITY con datos
    if quality_score >= 65:
        details = []
        if roe: details.append(f"ROE {roe*100:.1f}%" if roe < 1 else f"ROE {roe:.1f}%")
        if op_margin: details.append(f"margen {op_margin*100:.1f}%" if op_margin < 1 else f"margen {op_margin:.1f}%")
        strengths.append(f"alta calidad ({', '.join(details) if details else 'métricas top'})")
    elif quality_score <= 35:
        weaknesses.append(f"baja calidad operativa (ROE bajo, márgenes débiles)")

    # MOMENTUM con datos
    if momentum_score >= 65:
        mom_str = f"retorno 12M: {mom_12m:.1f}%" if mom_12m else "tendencia alcista"
        strengths.append(f"fuerte momentum ({mom_str})")
    elif momentum_score <= 35:
        mom_str = f"retorno 12M: {mom_12m:.1f}%" if mom_12m else "tendencia bajista"
        weaknesses.append(f"momentum débil ({mom_str})")

    # LOWVOL con datos
    if lowvol_score >= 65:
        vol_str = []
        if beta: vol_str.append(f"beta {beta:.2f}")
        if vol: vol_str.append(f"vol {vol:.1f}%")
        strengths.append(f"baja volatilidad ({', '.join(vol_str) if vol_str else 'defensivo'})")
    elif lowvol_score <= 35:
        vol_str = []
        if beta: vol_str.append(f"beta {beta:.2f}")
        if vol: vol_str.append(f"vol {vol:.1f}%")
        weaknesses.append(f"alta volatilidad ({', '.join(vol_str) if vol_str else 'riesgo elevado'})")

    # CONGRESS con detalles
    if congress_score > 60:
        if congress_details and congress_details.get('trades'):
            trades_info = []
            for trade in congress_details.get('recent_trades', [])[:2]:
                pol = trade.get('politician', 'Congresista')
                action = 'COMPRA' if 'PURCHASE' in str(trade.get('type', '')).upper() else 'VENTA'
                amount = trade.get('amount_range', '')
                date = trade.get('transaction_date', '')[:10] if trade.get('transaction_date') else ''
                trades_info.append(f"{pol} {action} {amount} ({date})")
            if trades_info:
                strengths.append(f"congresistas comprando: {'; '.join(trades_info)}")
            else:
                strengths.append("congresistas comprando recientemente")
        else:
            strengths.append("congresistas comprando recientemente")
    elif congress_score < 40:
        weaknesses.append("congresistas vendiendo")

    # POLYMARKET
    if poly_score > 60:
        strengths.append("smart money positivo (Polymarket)")
    elif poly_score < 40:
        weaknesses.append("smart money negativo (Polymarket)")

    # === CONSTRUIR TESIS COMPLETA ===
    parts = []

    # Header con señal y score
    signal_text = {
        'STRONG BUY': 'COMPRA FUERTE',
        'BUY': 'COMPRA',
        'ACCUMULATE': 'ACUMULAR',
        'HOLD': 'MANTENER',
        'REDUCE': 'REDUCIR',
        'SELL': 'VENDER',
    }.get(signal, signal)

    parts.append(f"{signal_text} {ticker}")
    parts.append(f"Score: {score:.0f}/100")

    if target and price:
        parts.append(f"Precio: ${price:.2f} → Target: ${target:.2f} (Upside: {upside:+.0f}%)")

    # Fortalezas
    if strengths:
        parts.append(f"FORTALEZAS: {'; '.join(strengths)}")

    # Debilidades
    if weaknesses:
        parts.append(f"DEBILIDADES: {'; '.join(weaknesses)}")

    # RÉGIMEN MONETARIO (si disponible)
    if monetary_data:
        regime = monetary_data.get('liquidity_regime', 'NEUTRAL')
        risk = monetary_data.get('risk_appetite', 'NEUTRAL')
        net_liq = monetary_data.get('net_liquidity_change_30d', 0)

        # Determinar impacto según tipo de acción
        sector = c.get('sector', '')
        is_risk_on = ticker in ['NVDA', 'TSLA', 'AMD', 'COIN', 'NIO'] or sector in ['Technology']

        if regime == 'ABUNDANT_LIQUIDITY':
            if is_risk_on:
                parts.append(f"MACRO: Liquidez abundante (${net_liq:.0f}B 30d) FAVORECE este tipo de activo risk-on")
            else:
                parts.append(f"MACRO: Liquidez abundante (${net_liq:.0f}B 30d), entorno favorable para RV")
        elif regime == 'TIGHT_LIQUIDITY':
            if is_risk_on:
                parts.append(f"MACRO: Liquidez restrictiva (${net_liq:.0f}B 30d) PERJUDICA activos risk-on. PRECAUCIÓN")
            else:
                parts.append(f"MACRO: Liquidez restrictiva, favorecer defensivos")

    # Estrategia recomendada
    if signal in ['STRONG BUY', 'BUY'] and momentum_score > 65:
        parts.append("ESTRATEGIA: Corto plazo (trading) - momentum fuerte + catalizador")
    elif signal in ['STRONG BUY', 'BUY'] and quality_score > 60:
        parts.append("ESTRATEGIA: Medio plazo (swing) - calidad + valoración")
    elif quality_score > 70 and value_score > 55:
        parts.append("ESTRATEGIA: Largo plazo (value-quality) - empresa dominante a buen precio")
    elif signal == 'ACCUMULATE':
        parts.append("ESTRATEGIA: Ir acumulando gradualmente en debilidad")

    # Stop loss sugerido
    if signal in ['STRONG BUY', 'BUY'] and price:
        stop = price * 0.92  # 8% stop
        target_1 = price * 1.15  # 15% target
        parts.append(f"GESTIÓN RIESGO: Stop ${stop:.2f} (-8%), Target 1 ${target_1:.2f} (+15%)")

    return ' | '.join(parts)


def calculate_scores(companies):
    """Calcula scores multifactor para todas las empresas (V13: 6 factores)"""
    weights = get_final_weights()

    # V13: Inicializar trackers si están disponibles
    congress_tracker = None
    polymarket_client = None
    congress_signals = {}
    polymarket_signals = {}

    congress_details = {}  # Detalles completos para tesis
    if CONGRESS_AVAILABLE:
        try:
            congress_tracker = CongressTracker()
            print("  📊 Obteniendo señales de Congress...")
            for c in companies:
                ticker = c['ticker']
                signal = congress_tracker.get_signal_for_ticker(ticker, days=30)
                congress_signals[ticker] = signal.get('score', 50)
                # Guardar detalles completos para la tesis
                congress_details[ticker] = signal
        except Exception as e:
            print(f"  ⚠️ Congress tracker error: {e}")

    if POLYMARKET_AVAILABLE:
        try:
            polymarket_client = PolymarketClient()
            print("  📊 Obteniendo señales de Polymarket...")
            for c in companies:
                ticker = c['ticker']
                signal = polymarket_client.get_signal_for_ticker(ticker)
                polymarket_signals[ticker] = signal.get('score', 50)
        except Exception as e:
            print(f"  ⚠️ Polymarket client error: {e}")

    # Extraer métricas
    fwd_pes = [c.get('fwd_pe') or c.get('trailing_pe') for c in companies]
    ev_ebitdas = [c.get('ev_ebitda') for c in companies]
    fcf_yields = [c.get('fcf_yield') for c in companies]
    roics = [c.get('roic') or c.get('roe') for c in companies]
    op_margins = [c.get('op_margin') for c in companies]
    net_debt_ebitdas = [c.get('net_debt_ebitda') for c in companies]
    betas = [c.get('beta') for c in companies]
    vols = [c.get('vol_252d') for c in companies]
    mom_12ms = [c.get('mom_12m') for c in companies]
    analyst_revs = [c.get('analyst_rev') for c in companies]

    # Calcular percentiles
    value_pe_pctl = percentile_rank(fwd_pes, reverse=True)
    value_ev_pctl = percentile_rank(ev_ebitdas, reverse=True)
    value_fcf_pctl = percentile_rank(fcf_yields, reverse=False)
    qual_roic_pctl = percentile_rank(roics, reverse=False)
    qual_margin_pctl = percentile_rank(op_margins, reverse=False)
    qual_debt_pctl = percentile_rank(net_debt_ebitdas, reverse=True)
    mom_return_pctl = percentile_rank(mom_12ms, reverse=False)
    mom_rev_pctl = percentile_rank(analyst_revs, reverse=False)
    risk_beta_pctl = percentile_rank(betas, reverse=True)
    risk_vol_pctl = percentile_rank(vols, reverse=True)

    for i, c in enumerate(companies):
        # Factor scores tradicionales
        c['value_score'] = (value_pe_pctl[i] + value_ev_pctl[i] + value_fcf_pctl[i]) / 3
        c['quality_score'] = (qual_roic_pctl[i] + qual_margin_pctl[i] + qual_debt_pctl[i]) / 3
        c['momentum_score'] = (mom_return_pctl[i] + mom_rev_pctl[i]) / 2
        c['lowvol_score'] = (risk_beta_pctl[i] + risk_vol_pctl[i]) / 2

        # V13: Nuevos factores
        ticker = c['ticker']
        c['congress_score'] = congress_signals.get(ticker, 50)
        c['polymarket_score'] = polymarket_signals.get(ticker, 50)
        c['news_score'] = 50  # Reservado para futuro uso

        # Ajuste regional
        region_adj = 0
        if c['country'] == 'China':
            region_adj = -10
        elif c['region'] == 'Europe' and c['sector'] in ['Defensa', 'Energia']:
            region_adj = +5
        c['region_adj'] = region_adj

        # Score compuesto V13 (6 factores + ajustes)
        c['composite_score'] = (
            c['value_score'] * weights.get('value', 0.20) +
            c['quality_score'] * weights.get('quality', 0.25) +
            c['momentum_score'] * weights.get('momentum', 0.15) +
            c['lowvol_score'] * weights.get('lowvol', 0.15) +
            c['congress_score'] * weights.get('congress', 0.10) +
            c['polymarket_score'] * weights.get('polymarket', 0.10) +
            region_adj
        )

        # Drivers (incluye nuevos factores)
        factor_contributions = {
            'Value': (c['value_score'] - 50) * weights.get('value', 0.20),
            'Quality': (c['quality_score'] - 50) * weights.get('quality', 0.25),
            'Momentum': (c['momentum_score'] - 50) * weights.get('momentum', 0.15),
            'LowVol': (c['lowvol_score'] - 50) * weights.get('lowvol', 0.15),
            'Congress': (c['congress_score'] - 50) * weights.get('congress', 0.10),
            'Polymarket': (c['polymarket_score'] - 50) * weights.get('polymarket', 0.10),
        }
        sorted_factors = sorted(factor_contributions.items(),
                               key=lambda x: abs(x[1]), reverse=True)

        positive = [(f, round(v, 1)) for f, v in sorted_factors if v > 0][:3]
        negative = [(f, round(v, 1)) for f, v in sorted_factors if v < 0][:3]

        c['drivers_str'] = ' + '.join([f"{f}({v:+.1f})" for f, v in positive]) if positive else "None"
        c['detractors_str'] = ' - '.join([f"{f}({abs(v):.1f})" for f, v in negative]) if negative else "None"

        # Upside PRIMERO (necesario para la señal)
        if c['target'] and c['price']:
            c['upside'] = ((c['target'] - c['price']) / c['price']) * 100
        else:
            c['upside'] = 0

        # Señal - CORREGIDA: combina score + upside
        # No puede ser BUY si el upside es negativo
        score = c['composite_score']
        upside = c['upside']

        if score >= 60 and upside >= 10:
            c['signal'] = 'STRONG BUY'
        elif score >= 60 and upside >= 0:
            c['signal'] = 'BUY'
        elif score >= 60 and upside < 0:
            c['signal'] = 'HOLD'  # Score alto pero upside negativo = HOLD, no BUY
        elif score >= 50 and upside >= 0:
            c['signal'] = 'ACCUMULATE'
        elif score >= 50 and upside < 0:
            c['signal'] = 'HOLD'
        elif score >= 40:
            c['signal'] = 'HOLD'
        elif score >= 30:
            c['signal'] = 'REDUCE'
        else:
            c['signal'] = 'SELL'

        # Advertencia si hay discrepancia score vs upside
        if score >= 55 and upside < -10:
            c['warning'] = 'SCORE_UPSIDE_MISMATCH'
        elif score < 45 and upside > 20:
            c['warning'] = 'UPSIDE_SCORE_MISMATCH'
        else:
            c['warning'] = ''

        # Generar tesis de inversión profesional CON datos de monetary y congress
        ticker = c['ticker']
        cong_detail = congress_details.get(ticker, {})
        c['investment_thesis'] = generate_investment_thesis(c, monetary_data=None, congress_details=cong_detail)
        c['congress_details'] = cong_detail  # Guardar para Excel

    # Ordenar por score
    companies.sort(key=lambda x: x['composite_score'], reverse=True)
    return companies


# =============================================================================
# GENERACIÓN DE EXCEL
# =============================================================================

def create_excel(companies, output_path):
    """Genera el Excel con todas las hojas"""
    wb = Workbook()
    weights = get_final_weights()
    
    # =========================================================================
    # HOJA 1: DASHBOARD (Mejorado con ratios y tabla ordenable)
    # =========================================================================
    ws_dash = wb.active
    ws_dash.title = 'Dashboard'

    # Obtener info de monetary plumbing para mostrar contexto
    monetary_regime = 'N/A'
    monetary_impact = ''
    if MONETARY_ANALYSIS:
        monetary_regime = MONETARY_ANALYSIS.get('regime', 'NEUTRAL')
        crisis_signals = MONETARY_ANALYSIS.get('crisis_signals', 0)
        if monetary_regime == 'ABUNDANT_LIQUIDITY':
            monetary_impact = 'Liquidez alta: +Momentum, -Defensivo'
        elif monetary_regime == 'TIGHT_LIQUIDITY':
            monetary_impact = 'Liquidez baja: +Quality/LowVol, -Momentum'
        elif monetary_regime == 'CRISIS_MODE':
            monetary_impact = 'CRISIS: Max defensivo, min riesgo'
        else:
            monetary_impact = 'Neutral: pesos base'

    ws_dash['A1'] = f'MULTIFACTOR DASHBOARD - {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws_dash['A1'].font = title_font
    ws_dash.merge_cells('A1:R1')

    # Línea 2: Macro context con monetary
    ws_dash['A2'] = f'Macro: {MACRO_CONFIG["regime"]} | Liquidez: {monetary_regime} | {monetary_impact}'
    ws_dash['A2'].font = Font(italic=True, size=10)
    if monetary_regime == 'CRISIS_MODE':
        ws_dash['A2'].fill = red_fill
    elif monetary_regime == 'TIGHT_LIQUIDITY':
        ws_dash['A2'].fill = yellow_fill
    elif monetary_regime == 'ABUNDANT_LIQUIDITY':
        ws_dash['A2'].fill = green_fill

    # Línea 3: Pesos actuales
    ws_dash['A3'] = f'Pesos: Value {weights.get("value",0.20)*100:.0f}% | Quality {weights.get("quality",0.25)*100:.0f}% | Momentum {weights.get("momentum",0.15)*100:.0f}% | LowVol {weights.get("lowvol",0.15)*100:.0f}% | Congress {weights.get("congress",0.10)*100:.0f}% | Poly {weights.get("polymarket",0.10)*100:.0f}%'
    ws_dash['A3'].font = Font(size=9, color='666666')

    # Headers mejorados con ratios fundamentales
    headers = ['Ticker', 'Empresa', 'Sector', 'Precio', 'Score', 'Senal',
               'P/E', 'EV/EBITDA', 'Deuda/EBITDA', 'ROE%', 'FCF_Yield%',
               'Val', 'Qual', 'Mom', 'LowVol', 'Cong', 'Poly', 'Tesis']

    header_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws_dash.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center

    for i, c in enumerate(companies, 1):
        row = i + header_row

        # Ticker y empresa
        ws_dash.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
        ws_dash.cell(row=row, column=2, value=c['name'])
        ws_dash.cell(row=row, column=3, value=c['sector'])

        # Precio
        cell_price = ws_dash.cell(row=row, column=4, value=c['price'])
        cell_price.number_format = '$#,##0.00'

        # Score compuesto - FORMULA EXCEL para transparencia
        # Columnas: L=Val, M=Qual, N=Mom, O=LowVol, P=Cong, Q=Poly
        w_val = weights.get('value', 0.20)
        w_qual = weights.get('quality', 0.25)
        w_mom = weights.get('momentum', 0.15)
        w_lvol = weights.get('lowvol', 0.15)
        w_cong = weights.get('congress', 0.10)
        w_poly = weights.get('polymarket', 0.10)

        # Formula: =L*w + M*w + N*w + O*w + P*w + Q*w
        score_formula = f"=ROUND(L{row}*{w_val}+M{row}*{w_qual}+N{row}*{w_mom}+O{row}*{w_lvol}+P{row}*{w_cong}+Q{row}*{w_poly},1)"
        cell_score = ws_dash.cell(row=row, column=5, value=score_formula)
        cell_score.alignment = center
        # Formato condicional basado en el valor calculado (para colores usamos el valor pre-calculado)
        if c['composite_score'] >= 60:
            cell_score.fill = green_fill
            cell_score.font = green_font
        elif c['composite_score'] >= 50:
            cell_score.fill = light_green_fill
        elif c['composite_score'] < 40:
            cell_score.fill = red_fill
            cell_score.font = red_font

        # Señal
        cell_signal = ws_dash.cell(row=row, column=6, value=c['signal'])
        cell_signal.alignment = center
        if c['signal'] in ['BUY', 'STRONG BUY']:
            cell_signal.fill = green_fill
            cell_signal.font = green_font
        elif c['signal'] == 'ACCUMULATE':
            cell_signal.fill = light_green_fill
        elif c['signal'] == 'HOLD':
            cell_signal.fill = yellow_fill
        elif c['signal'] in ['REDUCE', 'SELL']:
            cell_signal.fill = red_fill
            cell_signal.font = red_font

        # RATIOS FUNDAMENTALES (nuevos) - para entender el porqué
        # P/E (col 7)
        pe = c.get('fwd_pe') or c.get('trailing_pe')
        cell_pe = ws_dash.cell(row=row, column=7, value=round(pe, 1) if pe else '-')
        if pe and pe < 15:
            cell_pe.fill = green_fill
        elif pe and pe > 30:
            cell_pe.fill = red_fill

        # EV/EBITDA (col 8)
        ev_ebitda = c.get('ev_ebitda')
        cell_ev = ws_dash.cell(row=row, column=8, value=round(ev_ebitda, 1) if ev_ebitda else '-')
        if ev_ebitda and ev_ebitda < 10:
            cell_ev.fill = green_fill
        elif ev_ebitda and ev_ebitda > 20:
            cell_ev.fill = red_fill

        # Deuda/EBITDA (col 9) - ratio de apalancamiento clave
        debt_ebitda = c.get('net_debt_ebitda')
        cell_debt = ws_dash.cell(row=row, column=9, value=round(debt_ebitda, 1) if debt_ebitda else '-')
        if debt_ebitda:
            if debt_ebitda < 2:
                cell_debt.fill = green_fill
            elif debt_ebitda > 4:
                cell_debt.fill = red_fill

        # ROE% (col 10)
        roe = c.get('roe') or c.get('roic')
        cell_roe = ws_dash.cell(row=row, column=10, value=f"{roe*100:.0f}" if roe else '-')
        if roe and roe > 0.15:
            cell_roe.fill = green_fill
        elif roe and roe < 0.05:
            cell_roe.fill = red_fill

        # FCF Yield% (col 11)
        fcf = c.get('fcf_yield')
        cell_fcf = ws_dash.cell(row=row, column=11, value=f"{fcf*100:.1f}" if fcf else '-')
        if fcf and fcf > 0.05:
            cell_fcf.fill = green_fill
        elif fcf and fcf < 0:
            cell_fcf.fill = red_fill

        # Factor scores (12-17)
        ws_dash.cell(row=row, column=12, value=round(c['value_score'], 0)).alignment = center
        ws_dash.cell(row=row, column=13, value=round(c['quality_score'], 0)).alignment = center
        ws_dash.cell(row=row, column=14, value=round(c['momentum_score'], 0)).alignment = center
        ws_dash.cell(row=row, column=15, value=round(c.get('lowvol_score', 50), 0)).alignment = center

        # Congress score (16) con color
        cong_score = c.get('congress_score', 50)
        cell_cong = ws_dash.cell(row=row, column=16, value=round(cong_score, 0))
        cell_cong.alignment = center
        if cong_score >= 65:
            cell_cong.fill = green_fill
        elif cong_score <= 35:
            cell_cong.fill = red_fill

        # Polymarket score (17) con color
        poly_score = c.get('polymarket_score', 50)
        cell_poly = ws_dash.cell(row=row, column=17, value=round(poly_score, 0))
        cell_poly.alignment = center
        if poly_score >= 65:
            cell_poly.fill = green_fill
        elif poly_score <= 35:
            cell_poly.fill = red_fill

        # Tesis completa (18)
        thesis = c.get('investment_thesis', c.get('drivers_str', ''))
        cell_thesis = ws_dash.cell(row=row, column=18, value=thesis)
        cell_thesis.font = Font(size=8)
        cell_thesis.alignment = Alignment(wrap_text=True, vertical='top')

        # Borders para toda la fila
        for col in range(1, 19):
            ws_dash.cell(row=row, column=col).border = thin_border

    # Crear tabla Excel para ordenar/filtrar
    table_end_row = header_row + len(companies)
    table_ref = f"A{header_row}:R{table_end_row}"
    table = Table(displayName="DashboardTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws_dash.add_table(table)

    set_col_widths(ws_dash, [8, 18, 12, 10, 7, 10, 7, 9, 10, 7, 9, 6, 6, 6, 6, 6, 6, 100])

    # Leyenda de columnas al final
    legend_row = table_end_row + 3
    ws_dash.cell(row=legend_row, column=1, value='LEYENDA DE COLUMNAS:').font = Font(bold=True, size=11)
    legends = [
        ('Val (Value)', 'Valoracion: P/E bajo, EV/EBITDA bajo = mejor. Score alto = empresa barata.'),
        ('Qual (Quality)', 'Calidad: ROE alto, margenes altos, deuda baja = mejor. Score alto = empresa solida.'),
        ('Mom (Momentum)', 'Momentum: Precio subiendo vs 12m atras. Score alto = tendencia alcista fuerte.'),
        ('LowVol', 'Baja Volatilidad: Beta bajo, volatilidad historica baja. Score alto = menos riesgo.'),
        ('Cong (Congress)', 'Trades Congresistas: >50 = compras recientes, <50 = ventas. 50 = sin actividad.'),
        ('Poly (Polymarket)', 'Smart Money: Apuestas grandes en eventos que afectan la accion. >50 = bullish.'),
        ('Deuda/EBITDA', 'Apalancamiento: <2x = conservador (verde), 2-4x = normal, >4x = alto riesgo (rojo).'),
    ]
    for i, (term, desc) in enumerate(legends):
        ws_dash.cell(row=legend_row + 1 + i, column=1, value=term).font = Font(bold=True, size=9)
        ws_dash.cell(row=legend_row + 1 + i, column=2, value=desc).font = Font(size=9)
        ws_dash.merge_cells(start_row=legend_row + 1 + i, start_column=2, end_row=legend_row + 1 + i, end_column=10)
    
    # =========================================================================
    # HOJA 2: RAW DATA
    # =========================================================================
    ws_raw = wb.create_sheet('Raw_Data')
    
    raw_headers = ['Ticker', 'Name', 'Sector', 'Country', 'Price', 'Fwd_PE', 
                   'EV_EBITDA', 'FCF_Yield', 'ROIC', 'Op_Margin', 'Net_Debt_EBITDA',
                   'Beta', 'Vol_252d', 'Mom_12M', 'Target', 'Rating', 'Div_Yield']
    
    for col, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    
    for i, c in enumerate(companies, 1):
        row = i + 1
        ws_raw.cell(row=row, column=1, value=c['ticker'])
        ws_raw.cell(row=row, column=2, value=c['name'])
        ws_raw.cell(row=row, column=3, value=c['sector'])
        ws_raw.cell(row=row, column=4, value=c['country'])
        ws_raw.cell(row=row, column=5, value=c['price'])
        ws_raw.cell(row=row, column=6, value=c.get('fwd_pe'))
        ws_raw.cell(row=row, column=7, value=c.get('ev_ebitda'))
        ws_raw.cell(row=row, column=8, value=c.get('fcf_yield'))
        ws_raw.cell(row=row, column=9, value=c.get('roic'))
        ws_raw.cell(row=row, column=10, value=c.get('op_margin'))
        ws_raw.cell(row=row, column=11, value=c.get('net_debt_ebitda'))
        ws_raw.cell(row=row, column=12, value=c.get('beta'))
        ws_raw.cell(row=row, column=13, value=c.get('vol_252d'))
        ws_raw.cell(row=row, column=14, value=c.get('mom_12m'))
        ws_raw.cell(row=row, column=15, value=c.get('target'))
        ws_raw.cell(row=row, column=16, value=c.get('rating'))
        ws_raw.cell(row=row, column=17, value=c.get('dividend_yield'))
    
    set_col_widths(ws_raw, [8, 22, 14, 10, 10, 10, 12, 10, 10, 12, 14, 8, 10, 10, 10, 8, 10])
    
    # =========================================================================
    # HOJA 3: ACTION LIST
    # =========================================================================
    ws_action = wb.create_sheet('Action_List')
    
    ws_action['A1'] = f'🎯 ACTION LIST - {datetime.now().strftime("%d %b %Y")}'
    ws_action['A1'].font = title_font
    ws_action.merge_cells('A1:F1')
    
    # STRONG BUY + BUY
    ws_action['A3'] = 'COMPRAR (Score >=60 Y Upside positivo)'
    ws_action['A3'].font = Font(bold=True, size=12, color='006100')
    ws_action['A3'].fill = green_fill

    action_headers = ['Ticker', 'Empresa', 'Score', 'Precio', 'Target', 'Upside%', 'Senal', 'Drivers']
    for col, h in enumerate(action_headers, 1):
        cell = ws_action.cell(row=4, column=col, value=h)
        cell.fill = light_green_fill
        cell.font = Font(bold=True)

    row = 5
    buy_list = [c for c in companies if c['signal'] in ['STRONG BUY', 'BUY']]
    for c in buy_list:
        ws_action.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
        ws_action.cell(row=row, column=2, value=c['name'])
        ws_action.cell(row=row, column=3, value=round(c['composite_score'], 1))
        ws_action.cell(row=row, column=4, value=c['price'])
        ws_action.cell(row=row, column=5, value=c.get('target', ''))
        cell_upside = ws_action.cell(row=row, column=6, value=f"{c['upside']:.0f}%" if c['upside'] else '')
        if c['upside'] and c['upside'] >= 10:
            cell_upside.fill = green_fill
            cell_upside.font = green_font
        cell_sig = ws_action.cell(row=row, column=7, value=c['signal'])
        if c['signal'] == 'STRONG BUY':
            cell_sig.fill = green_fill
            cell_sig.font = green_font
        ws_action.cell(row=row, column=8, value=c['drivers_str'])
        for col in range(1, 9):
            ws_action.cell(row=row, column=col).border = thin_border
        row += 1

    if not buy_list:
        ws_action.cell(row=row, column=1, value='No hay senales BUY (requiere Score>=60 Y Upside>=0%)')
        row += 1
    
    # ACCUMULATE
    row += 1
    ws_action.cell(row=row, column=1, value='ACUMULAR (Score 50-59 Y Upside positivo)').font = Font(bold=True, size=12)
    ws_action.cell(row=row, column=1).fill = yellow_fill
    row += 1

    acc_headers = ['Ticker', 'Empresa', 'Score', 'Precio', 'Target', 'Upside%', 'Drivers']
    for col, h in enumerate(acc_headers, 1):
        cell = ws_action.cell(row=row, column=col, value=h)
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
    row += 1

    acc_list = [c for c in companies if c['signal'] == 'ACCUMULATE'][:10]
    for c in acc_list:
        ws_action.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
        ws_action.cell(row=row, column=2, value=c['name'])
        ws_action.cell(row=row, column=3, value=round(c['composite_score'], 1))
        ws_action.cell(row=row, column=4, value=c['price'])
        ws_action.cell(row=row, column=5, value=c.get('target', ''))
        cell_upside = ws_action.cell(row=row, column=6, value=f"{c['upside']:.0f}%" if c['upside'] else '')
        if c['upside'] and c['upside'] >= 5:
            cell_upside.fill = light_green_fill
        ws_action.cell(row=row, column=7, value=c['drivers_str'])
        for col in range(1, 8):
            ws_action.cell(row=row, column=col).border = thin_border
        row += 1
    
    # AVOID
    row += 1
    ws_action.cell(row=row, column=1, value='🔴 EVITAR (Score <40)').font = Font(bold=True, size=12, color='9C0006')
    ws_action.cell(row=row, column=1).fill = red_fill
    row += 1
    
    avoid_headers = ['Ticker', 'Empresa', 'Score', 'Señal', 'Detractors']
    for col, h in enumerate(avoid_headers, 1):
        cell = ws_action.cell(row=row, column=col, value=h)
        cell.fill = red_fill
        cell.font = Font(bold=True, color='FFFFFF')
    row += 1
    
    avoid_list = [c for c in companies if c['signal'] in ['REDUCE', 'SELL']]
    for c in avoid_list:
        ws_action.cell(row=row, column=1, value=c['ticker'])
        ws_action.cell(row=row, column=2, value=c['name'])
        ws_action.cell(row=row, column=3, value=round(c['composite_score'], 1))
        ws_action.cell(row=row, column=4, value=c['signal'])
        ws_action.cell(row=row, column=5, value=c['detractors_str'])
        for col in range(1, 6):
            ws_action.cell(row=row, column=col).border = thin_border
        row += 1
    
    set_col_widths(ws_action, [10, 22, 8, 10, 10, 10, 12, 40])

    # =========================================================================
    # HOJA 6: SCORE BREAKDOWN (Glosario movido al final)
    # =========================================================================
    ws_breakdown = wb.create_sheet('Score_Breakdown')

    ws_breakdown['A1'] = 'DESGLOSE DE SCORES - Como se calcula cada puntuacion'
    ws_breakdown['A1'].font = title_font
    ws_breakdown.merge_cells('A1:M1')

    breakdown_headers = ['Ticker', 'Precio', 'Target', 'Upside%',
                         'Value', 'Quality', 'Momentum', 'LowVol', 'Congress', 'Polymarket',
                         'SCORE', 'Senal', 'Tesis de Inversion']

    for col, h in enumerate(breakdown_headers, 1):
        cell = ws_breakdown.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center

    # Subheaders explicando los pesos
    ws_breakdown['A2'] = f'Pesos actuales: Value {weights.get("value",0.20)*100:.0f}% | Quality {weights.get("quality",0.25)*100:.0f}% | Mom {weights.get("momentum",0.15)*100:.0f}% | LowVol {weights.get("lowvol",0.15)*100:.0f}% | Congress {weights.get("congress",0.10)*100:.0f}% | Polymarket {weights.get("polymarket",0.10)*100:.0f}%'
    ws_breakdown['A2'].font = Font(italic=True, size=10)

    for i, c in enumerate(companies, 1):
        row = i + 3
        ws_breakdown.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
        ws_breakdown.cell(row=row, column=2, value=c['price'])
        ws_breakdown.cell(row=row, column=3, value=c.get('target', ''))

        # Upside con color
        cell_up = ws_breakdown.cell(row=row, column=4, value=round(c['upside'], 1) if c['upside'] else 0)
        if c['upside'] and c['upside'] >= 10:
            cell_up.fill = green_fill
            cell_up.font = green_font
        elif c['upside'] and c['upside'] < 0:
            cell_up.fill = red_fill
            cell_up.font = red_font

        # Factor scores individuales
        for col_idx, score_key in enumerate(['value_score', 'quality_score', 'momentum_score',
                                              'lowvol_score', 'congress_score', 'polymarket_score'], 5):
            score_val = c.get(score_key, 50)
            cell = ws_breakdown.cell(row=row, column=col_idx, value=round(score_val, 0))
            cell.alignment = center
            # Colorear segun percentil
            if score_val >= 70:
                cell.fill = green_fill
            elif score_val >= 60:
                cell.fill = light_green_fill
            elif score_val <= 30:
                cell.fill = red_fill
            elif score_val <= 40:
                cell.fill = yellow_fill

        # Score compuesto CON FÓRMULA EXCEL (para que el usuario vea cómo se calcula)
        # Fórmula: Value*20% + Quality*25% + Momentum*15% + LowVol*15% + Congress*10% + Poly*10%
        w_val = weights.get('value', 0.20)
        w_qual = weights.get('quality', 0.25)
        w_mom = weights.get('momentum', 0.15)
        w_vol = weights.get('lowvol', 0.15)
        w_cong = weights.get('congress', 0.10)
        w_poly = weights.get('polymarket', 0.10)

        # Crear fórmula Excel visible
        formula = f"=E{row}*{w_val}+F{row}*{w_qual}+G{row}*{w_mom}+H{row}*{w_vol}+I{row}*{w_cong}+J{row}*{w_poly}"
        cell_score = ws_breakdown.cell(row=row, column=11, value=formula)
        cell_score.font = Font(bold=True)
        cell_score.alignment = center
        cell_score.number_format = '0.0'

        # Colorear según valor calculado (usamos el valor ya calculado en Python)
        if c['composite_score'] >= 60:
            cell_score.fill = green_fill
        elif c['composite_score'] < 40:
            cell_score.fill = red_fill

        # Señal
        cell_sig = ws_breakdown.cell(row=row, column=12, value=c['signal'])
        cell_sig.alignment = center
        if 'BUY' in c['signal']:
            cell_sig.fill = green_fill
        elif c['signal'] == 'HOLD':
            cell_sig.fill = yellow_fill
        elif c['signal'] in ['REDUCE', 'SELL']:
            cell_sig.fill = red_fill

        # Tesis de inversión (explica el porqué de la recomendación)
        thesis = c.get('investment_thesis', '')
        cell_thesis = ws_breakdown.cell(row=row, column=13, value=thesis)
        cell_thesis.font = Font(size=9)
        # Color según señal
        if 'COMPRA' in thesis:
            cell_thesis.font = Font(size=9, color='006100')
        elif 'VENDER' in thesis or 'REDUCIR' in thesis:
            cell_thesis.font = Font(size=9, color='9C0006')

        for col in range(1, 14):
            ws_breakdown.cell(row=row, column=col).border = thin_border

    set_col_widths(ws_breakdown, [8, 10, 10, 10, 8, 8, 10, 8, 10, 12, 8, 12, 60])

    # =========================================================================
    # HOJA 7: SPECULATIVE ANALYSIS (Mejorado con todos los trades de Congress)
    # =========================================================================
    ws_spec = wb.create_sheet('Speculative_Analysis')

    ws_spec['A1'] = 'CONGRESS TRADES - Historial Completo de Operaciones'
    ws_spec['A1'].font = title_font
    ws_spec.merge_cells('A1:H1')

    ws_spec['A2'] = 'Congresistas de EEUU deben reportar sus trades. Pelosi, Tuberville y otros tienen track records notables.'
    ws_spec['A2'].font = Font(italic=True, size=10)
    ws_spec['A3'] = 'Links: housestockwatcher.com (House) | senatestockwatcher.com (Senate)'
    ws_spec['A3'].font = Font(size=9, color='0000FF')

    # Obtener TODOS los trades recientes de la base de datos local
    all_congress_trades = []
    if SIGNAL_DB_AVAILABLE:
        try:
            db = SignalDatabase()
            # Obtener trades de los últimos 90 días
            all_congress_trades = db.get_recent_congress_trades(days=90)
        except Exception as e:
            print(f"  Error obteniendo trades de DB: {e}")

    # Si no hay datos de DB, intentar obtener de TRIGGERED_STOCKS
    if not all_congress_trades and TRIGGER_SUMMARY:
        for ticker_data in TRIGGER_SUMMARY:
            for t in ticker_data.get('triggers', []):
                if t.get('type') == 'CONGRESS_TRADE':
                    all_congress_trades.append({
                        'ticker': ticker_data.get('ticker', ''),
                        'politician': t.get('politician', 'N/A'),
                        'action': t.get('action', ''),
                        'transaction_date': t.get('date', ''),
                        'amount_range': t.get('amount', ''),
                        'chamber': t.get('chamber', ''),
                    })

    # Headers para la tabla de Congress
    row = 5
    ws_spec.cell(row=row, column=1, value='TRADES RECIENTES DE CONGRESISTAS (ultimos 90 dias)').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_spec.cell(row=row, column=1).fill = light_blue_fill
    row += 2

    congress_headers = ['Ticker', 'Politico', 'Partido', 'Accion', 'Monto', 'Fecha', 'Camara', 'Link']
    for col, h in enumerate(congress_headers, 1):
        cell = ws_spec.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    row += 1

    trades_shown = 0
    if all_congress_trades:
        # Ordenar por fecha más reciente
        try:
            all_congress_trades.sort(key=lambda x: x.get('transaction_date', ''), reverse=True)
        except:
            pass

        for trade in all_congress_trades[:50]:  # Mostrar hasta 50 trades
            ticker = trade.get('ticker', '')
            if not ticker or ticker == '--':
                continue

            ws_spec.cell(row=row, column=1, value=ticker).font = Font(bold=True)

            # Político
            politician = trade.get('politician', 'N/A')
            ws_spec.cell(row=row, column=2, value=politician)

            # Partido
            party = trade.get('party', '')
            cell_party = ws_spec.cell(row=row, column=3, value=party)
            if party == 'D':
                cell_party.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            elif party == 'R':
                cell_party.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

            # Acción (BUY/SELL)
            action = trade.get('type', trade.get('action', ''))
            if 'PURCHASE' in str(action).upper() or 'BUY' in str(action).upper():
                action_display = 'COMPRA'
                cell_action = ws_spec.cell(row=row, column=4, value=action_display)
                cell_action.fill = green_fill
                cell_action.font = green_font
            elif 'SALE' in str(action).upper() or 'SELL' in str(action).upper():
                action_display = 'VENTA'
                cell_action = ws_spec.cell(row=row, column=4, value=action_display)
                cell_action.fill = red_fill
                cell_action.font = red_font
            else:
                ws_spec.cell(row=row, column=4, value=action)

            # Monto
            amount = trade.get('amount_range', trade.get('amount', ''))
            ws_spec.cell(row=row, column=5, value=amount)

            # Fecha
            date = trade.get('transaction_date', '')
            ws_spec.cell(row=row, column=6, value=str(date)[:10] if date else '')

            # Cámara
            chamber = trade.get('chamber', '')
            ws_spec.cell(row=row, column=7, value=chamber)

            # Link (usar ticker-based URL que siempre funciona)
            link_url = f"https://housestockwatcher.com/summary_by_ticker/{ticker}"
            ws_spec.cell(row=row, column=8, value=link_url).font = Font(size=8, color='0000FF', underline='single')

            for col in range(1, 9):
                ws_spec.cell(row=row, column=col).border = thin_border
            row += 1
            trades_shown += 1

    if trades_shown == 0:
        ws_spec.cell(row=row, column=1, value='No hay trades de congresistas en la base de datos local.')
        ws_spec.cell(row=row, column=1).font = Font(italic=True)
        row += 1
        ws_spec.cell(row=row, column=1, value='Los datos se acumulan con cada ejecucion. Las APIs externas pueden estar temporalmente no disponibles.')
        row += 1

    # Sección de resumen por ticker (nuestro universo)
    row += 2
    ws_spec.cell(row=row, column=1, value='RESUMEN POR TICKER (nuestro universo)').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_spec.cell(row=row, column=1).fill = light_blue_fill
    row += 2

    summary_headers = ['Ticker', 'Congress Score', 'Interpretacion', 'Ultimo Trade']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_spec.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    row += 1

    for c in companies:
        cong_score = c.get('congress_score', 50)
        if cong_score != 50:  # Solo mostrar si hay señal
            ws_spec.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
            ws_spec.cell(row=row, column=2, value=round(cong_score, 0))

            if cong_score > 65:
                ws_spec.cell(row=row, column=3, value='BULLISH - Compras recientes')
                ws_spec.cell(row=row, column=3).fill = green_fill
            elif cong_score > 50:
                ws_spec.cell(row=row, column=3, value='Ligeramente bullish')
                ws_spec.cell(row=row, column=3).fill = light_green_fill
            elif cong_score < 35:
                ws_spec.cell(row=row, column=3, value='BEARISH - Ventas recientes')
                ws_spec.cell(row=row, column=3).fill = red_fill
            else:
                ws_spec.cell(row=row, column=3, value='Ligeramente bearish')

            # Buscar último trade para este ticker
            last_trade = ''
            for trade in all_congress_trades:
                if trade.get('ticker', '').upper() == c['ticker'].upper():
                    pol = trade.get('politician', '')[:15]
                    act = 'C' if 'PURCHASE' in str(trade.get('type', '')).upper() else 'V'
                    last_trade = f"{pol} ({act})"
                    break
            ws_spec.cell(row=row, column=4, value=last_trade)

            for col in range(1, 5):
                ws_spec.cell(row=row, column=col).border = thin_border
            row += 1

    # Sección Polymarket
    row += 2
    ws_spec.cell(row=row, column=1, value='POLYMARKET SMART MONEY').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_spec.cell(row=row, column=1).fill = light_blue_fill
    row += 1
    ws_spec.cell(row=row, column=1, value='Mercado de predicciones: apuestas >$50k pueden indicar informacion privilegiada.')
    ws_spec.cell(row=row, column=1).font = Font(italic=True, size=10)
    row += 2

    poly_headers = ['Ticker', 'Polymarket Score', 'Interpretacion']
    for col, h in enumerate(poly_headers, 1):
        cell = ws_spec.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    row += 1

    poly_found = False
    for c in companies:
        poly_score = c.get('polymarket_score', 50)
        if poly_score != 50:
            poly_found = True
            ws_spec.cell(row=row, column=1, value=c['ticker']).font = Font(bold=True)
            ws_spec.cell(row=row, column=2, value=round(poly_score, 0))
            if poly_score > 60:
                ws_spec.cell(row=row, column=3, value='BULLISH - Smart money detectado')
                ws_spec.cell(row=row, column=3).fill = green_fill
            elif poly_score < 40:
                ws_spec.cell(row=row, column=3, value='BEARISH - Smart money negativo')
                ws_spec.cell(row=row, column=3).fill = red_fill
            else:
                ws_spec.cell(row=row, column=3, value='NEUTRAL')
            for col in range(1, 4):
                ws_spec.cell(row=row, column=col).border = thin_border
            row += 1

    if not poly_found:
        ws_spec.cell(row=row, column=1, value='No hay senales de Polymarket actualmente')
        ws_spec.cell(row=row, column=1).font = Font(italic=True)

    set_col_widths(ws_spec, [10, 20, 8, 10, 18, 12, 10, 45])

    # =========================================================================
    # HOJA 8: BACKTEST SUMMARY
    # =========================================================================
    ws_bt = wb.create_sheet('Backtest_Summary')

    ws_bt['A1'] = 'RESUMEN DE BACKTEST - Validacion de Estrategias'
    ws_bt['A1'].font = title_font
    ws_bt.merge_cells('A1:E1')

    ws_bt['A3'] = 'Escenarios testeados: COVID Crash (2020), Recovery (2020), Bear (2022), AI Rally (2023), Market Highs (2024)'
    ws_bt['A3'].font = Font(italic=True)

    bt_data = [
        ['', '', '', '', ''],
        ['RESULTADOS POR CONFIGURACION DE PESOS', '', '', '', ''],
        ['Configuracion', 'Retorno Promedio', 'Sharpe Ratio', 'Win Rate', 'Alpha vs SPY'],
        ['V12 (4 factores)', '1.05%', '-4.27', '50.0%', '-6.06%'],
        ['V13 (6 factores)', '-0.27%', '-6.28', '48.0%', '-7.38%'],
        ['Momentum Heavy (MEJOR)', '33.47%', '-0.03', '57.4%', '+26.36%'],
        ['Defensive', '2.53%', '-2.03', '50.0%', '-4.58%'],
        ['Value Focused', '0.64%', '-2.17', '50.0%', '-6.48%'],
        ['', '', '', '', ''],
        ['CONCLUSIONES DEL BACKTEST', '', '', '', ''],
        ['', '', '', '', ''],
        ['1. Momentum es el factor mas rentable historicamente', '', '', '', ''],
        ['2. En mercados alcistas: mas peso a Momentum', '', '', '', ''],
        ['3. En mercados bajistas: Defensive (Quality + LowVol)', '', '', '', ''],
        ['4. Congress/Polymarket aportan valor en TIEMPO REAL (no backtesteable)', '', '', '', ''],
        ['', '', '', '', ''],
        ['PESOS RECOMENDADOS BASADOS EN BACKTEST', '', '', '', ''],
        ['Factor', 'Peso Actual', 'Peso Recomendado', 'Diferencia', ''],
        ['Value', '20%', '15%', '-5%', ''],
        ['Quality', '25%', '20%', '-5%', ''],
        ['Momentum', '15%', '35%', '+20%', 'SUBIR'],
        ['LowVol', '15%', '10%', '-5%', ''],
        ['Congress', '10%', '10%', '0%', ''],
        ['Polymarket', '10%', '10%', '0%', ''],
        ['', '', '', '', ''],
        ['NOTA: Los pesos actuales son conservadores.', '', '', '', ''],
        ['El backtest sugiere mas Momentum, pero eso aumenta el riesgo.', '', '', '', ''],
        ['Ajusta segun tu tolerancia al riesgo.', '', '', '', ''],
    ]

    for i, row_data in enumerate(bt_data):
        for col, val in enumerate(row_data, 1):
            cell = ws_bt.cell(row=i+5, column=col, value=val)
            if val and ('RESULTADOS' in val or 'CONCLUSIONES' in val or 'PESOS RECOMENDADOS' in val):
                cell.font = Font(bold=True, size=12, color=DARK_BLUE)
                cell.fill = light_blue_fill
            elif val in ['Configuracion', 'Retorno Promedio', 'Sharpe Ratio', 'Win Rate', 'Alpha vs SPY',
                        'Factor', 'Peso Actual', 'Peso Recomendado', 'Diferencia']:
                cell.fill = header_fill
                cell.font = white_font
            elif val and '(MEJOR)' in val:
                cell.fill = green_fill
                cell.font = green_font
            elif val and 'SUBIR' in val:
                cell.fill = green_fill
                cell.font = Font(bold=True, color='006100')
            cell.border = thin_border

    set_col_widths(ws_bt, [25, 18, 18, 15, 15])

    # =========================================================================
    # HOJA 9: LONG-TERM INVESTING (Mejorado con contexto monetario)
    # =========================================================================
    ws_lt = wb.create_sheet('Long_Term')

    ws_lt['A1'] = 'INVERSIONES A LARGO PLAZO - Value + Quality Focus'
    ws_lt['A1'].font = title_font
    ws_lt.merge_cells('A1:K1')

    ws_lt['A2'] = 'Estrategia: Empresas infravaloradas con alta calidad. Horizonte: 1-5 anos. Menor rotacion.'
    ws_lt['A2'].font = Font(italic=True)

    # CONTEXTO MONETARIO para largo plazo
    lt_monetary_context = ''
    if MONETARY_ANALYSIS:
        regime = MONETARY_ANALYSIS.get('regime', 'NEUTRAL')
        composite = MONETARY_ANALYSIS.get('composite_score', 50)
        vix_data = MONETARY_ANALYSIS.get('data', {}).get('vix', {})
        move_data = MONETARY_ANALYSIS.get('data', {}).get('move', {})

        if regime == 'ABUNDANT_LIQUIDITY':
            lt_monetary_context = f'CONTEXTO FAVORABLE: Liquidez abundante (score {composite:.0f}/100). Growth/Tech pueden funcionar bien a largo plazo.'
            ws_lt['A3'] = lt_monetary_context
            ws_lt['A3'].fill = green_fill
        elif regime == 'TIGHT_LIQUIDITY':
            lt_monetary_context = f'CONTEXTO CAUTELOSO: Liquidez restringida (score {composite:.0f}/100). Priorizar Quality y empresas con bajo endeudamiento.'
            ws_lt['A3'] = lt_monetary_context
            ws_lt['A3'].fill = yellow_fill
        elif regime == 'CRISIS_MODE':
            lt_monetary_context = f'CONTEXTO DE RIESGO: Senales de crisis activas. Maximo Quality/Defensivo, evitar deuda alta.'
            ws_lt['A3'] = lt_monetary_context
            ws_lt['A3'].fill = red_fill
        else:
            lt_monetary_context = f'CONTEXTO NEUTRAL: Condiciones mixtas (score {composite:.0f}/100). Diversificar entre Value y Quality.'
            ws_lt['A3'] = lt_monetary_context
    else:
        ws_lt['A3'] = 'Contexto monetario no disponible'

    ws_lt['A3'].font = Font(size=10, bold=True)

    # Calcular scores ponderados para largo plazo
    lt_weights = {'value': 0.35, 'quality': 0.35, 'momentum': 0.05, 'lowvol': 0.15, 'congress': 0.05, 'polymarket': 0.05}

    lt_companies = []
    for c in companies:
        lt_score = (
            c.get('value_score', 50) * lt_weights['value'] +
            c.get('quality_score', 50) * lt_weights['quality'] +
            c.get('momentum_score', 50) * lt_weights['momentum'] +
            c.get('lowvol_score', 50) * lt_weights['lowvol'] +
            c.get('congress_score', 50) * lt_weights['congress'] +
            c.get('polymarket_score', 50) * lt_weights['polymarket']
        )
        lt_companies.append({**c, 'lt_score': lt_score})

    lt_companies.sort(key=lambda x: x['lt_score'], reverse=True)

    ws_lt['A4'] = f'Pesos Long-Term: Value {lt_weights["value"]*100:.0f}% | Quality {lt_weights["quality"]*100:.0f}% | Mom {lt_weights["momentum"]*100:.0f}% | LowVol {lt_weights["lowvol"]*100:.0f}%'
    ws_lt['A4'].font = Font(size=9, color='666666')

    # Headers ampliados con contexto
    lt_headers = ['#', 'Ticker', 'Empresa', 'Sector', 'LT Score', 'Value', 'Quality', 'Deuda/EBITDA', 'Upside%', 'Senal LT', 'Por que']
    header_row = 6
    for col, h in enumerate(lt_headers, 1):
        cell = ws_lt.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font

    for i, c in enumerate(lt_companies[:20], 1):  # Top 20
        row = i + header_row
        ws_lt.cell(row=row, column=1, value=i)
        ws_lt.cell(row=row, column=2, value=c['ticker']).font = Font(bold=True)
        ws_lt.cell(row=row, column=3, value=c['name'])
        ws_lt.cell(row=row, column=4, value=c['sector'])

        cell_score = ws_lt.cell(row=row, column=5, value=round(c['lt_score'], 1))
        if c['lt_score'] >= 60:
            cell_score.fill = green_fill
        elif c['lt_score'] < 45:
            cell_score.fill = yellow_fill

        ws_lt.cell(row=row, column=6, value=round(c.get('value_score', 50), 0))
        ws_lt.cell(row=row, column=7, value=round(c.get('quality_score', 50), 0))

        # Deuda/EBITDA (nuevo)
        debt = c.get('net_debt_ebitda')
        cell_debt = ws_lt.cell(row=row, column=8, value=f"{debt:.1f}x" if debt else '-')
        if debt and debt < 2:
            cell_debt.fill = green_fill
        elif debt and debt > 4:
            cell_debt.fill = red_fill

        upside = c.get('upside', 0)
        cell_up = ws_lt.cell(row=row, column=9, value=f"{upside:.0f}%" if upside else '')
        if upside and upside >= 15:
            cell_up.fill = green_fill
        elif upside and upside < 0:
            cell_up.fill = red_fill

        # Señal para largo plazo
        if c['lt_score'] >= 60 and upside and upside >= 10:
            lt_signal = 'CORE HOLDING'
            ws_lt.cell(row=row, column=10, value=lt_signal).fill = green_fill
        elif c['lt_score'] >= 55 and upside and upside >= 0:
            lt_signal = 'ACCUMULATE'
            ws_lt.cell(row=row, column=10, value=lt_signal).fill = light_green_fill
        elif c['lt_score'] >= 50:
            lt_signal = 'WATCHLIST'
            ws_lt.cell(row=row, column=10, value=lt_signal)
        else:
            lt_signal = 'AVOID'
            ws_lt.cell(row=row, column=10, value=lt_signal).fill = yellow_fill

        # POR QUE - Explicación basada en scores y contexto monetario
        reasons = []
        if c.get('value_score', 50) >= 60:
            reasons.append('infravalorada')
        if c.get('quality_score', 50) >= 60:
            reasons.append('alta calidad')
        if debt and debt < 2:
            reasons.append('baja deuda')
        if c.get('momentum_score', 50) >= 60:
            reasons.append('tendencia alcista')

        # Añadir contexto monetario
        if MONETARY_ANALYSIS:
            regime = MONETARY_ANALYSIS.get('regime', 'NEUTRAL')
            sector = c.get('sector', '').lower()
            if regime == 'ABUNDANT_LIQUIDITY' and any(s in sector for s in ['tech', 'semi', 'ia', 'growth']):
                reasons.append('liquidez favorece tech')
            elif regime == 'TIGHT_LIQUIDITY' and c.get('quality_score', 50) >= 65:
                reasons.append('quality protege en tight')

        reason_text = ', '.join(reasons) if reasons else 'score general positivo'
        ws_lt.cell(row=row, column=11, value=reason_text.capitalize()).font = Font(size=9)

        for col in range(1, 12):
            ws_lt.cell(row=row, column=col).border = thin_border

    set_col_widths(ws_lt, [4, 8, 18, 12, 9, 7, 7, 10, 9, 12, 35])

    # =========================================================================
    # HOJA 10: MOMENTUM TRADING
    # =========================================================================
    ws_mom = wb.create_sheet('Momentum')

    ws_mom['A1'] = 'TRADING DE MOMENTUM - Corto/Medio Plazo'
    ws_mom['A1'].font = title_font
    ws_mom.merge_cells('A1:H1')

    ws_mom['A2'] = 'Estrategia: Seguir tendencias fuertes. Horizonte: 1-3 meses. Alta rotacion. SOLO en liquidez abundante.'
    ws_mom['A2'].font = Font(italic=True)

    # Calcular scores ponderados para momentum
    mom_weights = {'value': 0.05, 'quality': 0.10, 'momentum': 0.50, 'lowvol': 0.05, 'congress': 0.15, 'polymarket': 0.15}

    mom_companies = []
    for c in companies:
        mom_score = (
            c.get('value_score', 50) * mom_weights['value'] +
            c.get('quality_score', 50) * mom_weights['quality'] +
            c.get('momentum_score', 50) * mom_weights['momentum'] +
            c.get('lowvol_score', 50) * mom_weights['lowvol'] +
            c.get('congress_score', 50) * mom_weights['congress'] +
            c.get('polymarket_score', 50) * mom_weights['polymarket']
        )
        mom_companies.append({**c, 'mom_score': mom_score})

    mom_companies.sort(key=lambda x: x['mom_score'], reverse=True)

    ws_mom['A3'] = f'Pesos Momentum: Value {mom_weights["value"]*100:.0f}% | Quality {mom_weights["quality"]*100:.0f}% | Mom {mom_weights["momentum"]*100:.0f}% | Congress {mom_weights["congress"]*100:.0f}% | Polymarket {mom_weights["polymarket"]*100:.0f}%'
    ws_mom['A3'].font = Font(size=9, color='666666')

    # Warning sobre liquidez
    regime_warning = ''
    if MONETARY_ANALYSIS:
        regime = MONETARY_ANALYSIS.get('regime', 'UNKNOWN')
        if regime == 'TIGHT_LIQUIDITY':
            regime_warning = 'ATENCION: Liquidez restringida - REDUCIR exposicion a momentum'
            ws_mom['A4'] = regime_warning
            ws_mom['A4'].font = Font(bold=True, color='9C0006')
            ws_mom['A4'].fill = red_fill
        elif regime == 'ABUNDANT_LIQUIDITY':
            regime_warning = 'Liquidez abundante - Condiciones favorables para momentum'
            ws_mom['A4'] = regime_warning
            ws_mom['A4'].font = Font(bold=True, color='006100')
            ws_mom['A4'].fill = green_fill

    mom_headers = ['#', 'Ticker', 'Empresa', 'Mom Score', 'Mom 12M', 'Congress', 'Polymarket', 'Senal MOM']
    start_row = 6 if regime_warning else 5
    for col, h in enumerate(mom_headers, 1):
        cell = ws_mom.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font

    for i, c in enumerate(mom_companies[:15], 1):  # Top 15 momentum
        row = i + start_row
        ws_mom.cell(row=row, column=1, value=i)
        ws_mom.cell(row=row, column=2, value=c['ticker']).font = Font(bold=True)
        ws_mom.cell(row=row, column=3, value=c['name'])

        cell_score = ws_mom.cell(row=row, column=4, value=round(c['mom_score'], 1))
        if c['mom_score'] >= 60:
            cell_score.fill = green_fill

        mom_12m = c.get('mom_12m', 0)
        cell_mom = ws_mom.cell(row=row, column=5, value=f"{mom_12m:.0f}%" if mom_12m else '')
        if mom_12m and mom_12m >= 30:
            cell_mom.fill = green_fill
        elif mom_12m and mom_12m < 0:
            cell_mom.fill = red_fill

        ws_mom.cell(row=row, column=6, value=round(c.get('congress_score', 50), 0))
        ws_mom.cell(row=row, column=7, value=round(c.get('polymarket_score', 50), 0))

        # Senal para momentum
        if c['mom_score'] >= 65 and mom_12m and mom_12m > 20:
            mom_signal = 'STRONG MOM'
            ws_mom.cell(row=row, column=8, value=mom_signal).fill = green_fill
        elif c['mom_score'] >= 55:
            mom_signal = 'FOLLOW'
            ws_mom.cell(row=row, column=8, value=mom_signal).fill = light_green_fill
        else:
            mom_signal = 'WATCH'
            ws_mom.cell(row=row, column=8, value=mom_signal)

        for col in range(1, 9):
            ws_mom.cell(row=row, column=col).border = thin_border

    set_col_widths(ws_mom, [4, 8, 20, 10, 10, 10, 12, 12])

    # =========================================================================
    # HOJA 11: MONETARY PLUMBING (Mejorado con calculo detallado y MOVE)
    # =========================================================================
    ws_mp = wb.create_sheet('Monetary_Plumbing')

    ws_mp['A1'] = 'FONTANERIA MONETARIA - Calculo del Score de Liquidez'
    ws_mp['A1'].font = title_font
    ws_mp.merge_cells('A1:E1')

    ws_mp['A2'] = 'Analisis de liquidez macro que determina los pesos dinamicos del scoring'
    ws_mp['A2'].font = Font(italic=True)

    if MONETARY_ANALYSIS:
        ma = MONETARY_ANALYSIS

        # =================================================================
        # SECCION 1: REGIMEN Y SCORE ACTUAL
        # =================================================================
        ws_mp['A4'] = 'REGIMEN ACTUAL Y SCORE'
        ws_mp['A4'].font = Font(bold=True, size=12, color=DARK_BLUE)
        ws_mp['A4'].fill = light_blue_fill

        regime = ma.get('regime', 'UNKNOWN')
        score = ma.get('composite_score', 50)

        ws_mp['A5'] = f"Regimen: {regime}"
        ws_mp['A5'].font = Font(bold=True, size=14)
        if 'ABUNDANT' in regime:
            ws_mp['A5'].fill = green_fill
        elif 'TIGHT' in regime or 'CRISIS' in regime:
            ws_mp['A5'].fill = red_fill
        else:
            ws_mp['A5'].fill = yellow_fill

        ws_mp['A6'] = f"Score Liquidez: {score}/100"
        ws_mp['A6'].font = Font(bold=True, size=12)
        ws_mp['A7'] = ma.get('description', '')
        ws_mp['A8'] = f"Recomendacion: {ma.get('recommendation', '')}"
        ws_mp['A8'].font = Font(bold=True)

        # =================================================================
        # SECCION 2: CALCULO DEL SCORE (DESGLOSADO)
        # =================================================================
        row = 10
        ws_mp.cell(row=row, column=1, value='CALCULO DEL SCORE DE LIQUIDEZ').font = Font(bold=True, size=12, color=DARK_BLUE)
        ws_mp.cell(row=row, column=1).fill = light_blue_fill
        row += 2

        ws_mp.cell(row=row, column=1, value='Formula:').font = Font(bold=True)
        ws_mp.cell(row=row, column=2, value='Score = 0.30*NetLiq + 0.25*VIX + 0.20*Credit + 0.15*MOVE + 0.10*Japan')
        row += 2

        # Componentes individuales con su contribucion
        calc_headers = ['Componente', 'Valor', 'Condicion', 'Peso', 'Contribucion']
        for col, h in enumerate(calc_headers, 1):
            cell = ws_mp.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = white_font
        row += 1

        net_liq = ma.get('data', {}).get('net_liquidity', {})
        vix = ma.get('data', {}).get('vix', {})
        credit = ma.get('data', {}).get('credit_spreads', {})
        move = ma.get('data', {}).get('move', {})
        japan = ma.get('data', {}).get('japan', {})

        # Net Liquidity
        net_liq_score = net_liq.get('score', 50)
        ws_mp.cell(row=row, column=1, value='Net Liquidity')
        ws_mp.cell(row=row, column=2, value=f"${net_liq.get('net_liquidity_T', 'N/A')}T ({net_liq.get('change_1m_pct', 0):+.1f}%)")
        ws_mp.cell(row=row, column=3, value=net_liq.get('regime', 'N/A'))
        ws_mp.cell(row=row, column=4, value='30%')
        ws_mp.cell(row=row, column=5, value=f"{net_liq_score * 0.30:.1f}")
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 1

        # VIX
        vix_score = vix.get('score', 50)
        ws_mp.cell(row=row, column=1, value='VIX')
        ws_mp.cell(row=row, column=2, value=f"{vix.get('current', 'N/A'):.1f}" if isinstance(vix.get('current'), (int, float)) else 'N/A')
        ws_mp.cell(row=row, column=3, value=vix.get('condition', 'N/A'))
        ws_mp.cell(row=row, column=4, value='25%')
        ws_mp.cell(row=row, column=5, value=f"{vix_score * 0.25:.1f}")
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 1

        # Credit Spreads
        credit_score = credit.get('score', 50)
        ws_mp.cell(row=row, column=1, value='Credit Spreads')
        ws_mp.cell(row=row, column=2, value=f"{credit.get('spread_bps', 'N/A')} bps")
        ws_mp.cell(row=row, column=3, value=credit.get('condition', 'N/A'))
        ws_mp.cell(row=row, column=4, value='20%')
        ws_mp.cell(row=row, column=5, value=f"{credit_score * 0.20:.1f}")
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 1

        # MOVE (explicacion especial)
        move_score = move.get('score', 50) if move else 50
        move_current = move.get('current', 'N/A') if move else 'N/A'
        move_condition = move.get('condition', 'N/A') if move else 'N/A'
        ws_mp.cell(row=row, column=1, value='MOVE Index (proxy TLT)')
        ws_mp.cell(row=row, column=2, value=f"{move_current:.0f}" if isinstance(move_current, (int, float)) else 'N/A')
        ws_mp.cell(row=row, column=3, value=move_condition)
        ws_mp.cell(row=row, column=4, value='15%')
        ws_mp.cell(row=row, column=5, value=f"{move_score * 0.15:.1f}")
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 1

        # Japan
        japan_score = japan.get('score', 50) if japan else 50
        usdjpy = japan.get('usdjpy', {}).get('current', 'N/A') if japan else 'N/A'
        japan_condition = japan.get('condition', 'N/A') if japan else 'N/A'
        ws_mp.cell(row=row, column=1, value='Japan (USD/JPY + Banks)')
        ws_mp.cell(row=row, column=2, value=f"{usdjpy:.1f}" if isinstance(usdjpy, (int, float)) else 'N/A')
        ws_mp.cell(row=row, column=3, value=japan_condition)
        ws_mp.cell(row=row, column=4, value='10%')
        ws_mp.cell(row=row, column=5, value=f"{japan_score * 0.10:.1f}")
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 1

        # TOTAL
        ws_mp.cell(row=row, column=1, value='TOTAL SCORE').font = Font(bold=True)
        ws_mp.cell(row=row, column=5, value=f"{score:.1f}").font = Font(bold=True, size=12)
        ws_mp.cell(row=row, column=5).fill = green_fill if score >= 60 else (red_fill if score < 40 else yellow_fill)
        for c in range(1, 6):
            ws_mp.cell(row=row, column=c).border = thin_border
        row += 2

        # =================================================================
        # SECCION 3: MOVE INDEX - POR QUE NO LO USAMOS DIRECTAMENTE
        # =================================================================
        ws_mp.cell(row=row, column=1, value='SOBRE EL MOVE INDEX').font = Font(bold=True, size=12, color=DARK_BLUE)
        ws_mp.cell(row=row, column=1).fill = light_blue_fill
        row += 2

        move_explanation = [
            'El MOVE Index (Merrill Lynch Option Volatility Estimate) mide la volatilidad implicita de bonos del Tesoro.',
            'Es el "VIX de los bonos" - cuando sube, indica stress en renta fija que puede contagiar a acciones.',
            '',
            'POR QUE USAMOS UN PROXY (TLT volatilidad):',
            '- El MOVE real no esta disponible gratis en APIs publicas',
            '- Usamos la volatilidad de TLT (ETF de bonos largo plazo) * factor de escala',
            '- Correlacion historica con MOVE real: >0.85',
            '',
            'INTERPRETACION:',
            '- MOVE < 80: Mercado de bonos tranquilo, favorable para risk-on',
            '- MOVE 80-100: Volatilidad normal',
            '- MOVE 100-120: Volatilidad elevada, precaucion',
            '- MOVE > 120: Stress en bonos, riesgo de contagio a acciones',
            '- MOVE > 140: Crisis, reduce exposicion significativamente',
        ]

        for line in move_explanation:
            if line.startswith('POR QUE') or line.startswith('INTERPRETACION'):
                ws_mp.cell(row=row, column=1, value=line).font = Font(bold=True)
            else:
                ws_mp.cell(row=row, column=1, value=line)
            row += 1

        row += 1

        # =================================================================
        # SECCION 4: CONCLUSION Y OUTLOOK
        # =================================================================
        ws_mp.cell(row=row, column=1, value='MI CONCLUSION DE LIQUIDEZ ACTUAL').font = Font(bold=True, size=12, color=DARK_BLUE)
        ws_mp.cell(row=row, column=1).fill = light_blue_fill
        row += 2

        # Generar conclusion dinamica basada en los datos
        conclusion_lines = []
        if score >= 65:
            conclusion_lines.append('SITUACION ACTUAL: Liquidez ABUNDANTE - Condiciones favorables para risk assets.')
            conclusion_lines.append('El entorno apoya: Tech, Growth, Small Caps, Momentum fuerte.')
        elif score >= 50:
            conclusion_lines.append('SITUACION ACTUAL: Liquidez NEUTRAL - Condiciones mixtas.')
            conclusion_lines.append('Equilibrar entre Quality y Momentum. No hay urgencia pero tampoco euforia.')
        elif score >= 35:
            conclusion_lines.append('SITUACION ACTUAL: Liquidez RESTRINGIDA - Precaucion necesaria.')
            conclusion_lines.append('Favorecer Quality, Low Vol, Defensivos. Reducir exposicion a Growth agresivo.')
        else:
            conclusion_lines.append('SITUACION ACTUAL: CRISIS DE LIQUIDEZ - Modo defensivo.')
            conclusion_lines.append('Maxima precaucion. Cash, defensivos, cobertura. Evitar apalancamiento.')

        conclusion_lines.append('')
        conclusion_lines.append('ESTAR ATENTOS A:')

        # Alertas especificas
        if vix.get('current', 15) > 20:
            conclusion_lines.append(f'- VIX elevado ({vix.get("current", 0):.1f}): Volatilidad arriba de lo normal')
        if isinstance(move_current, (int, float)) and move_current > 100:
            conclusion_lines.append(f'- MOVE elevado ({move_current:.0f}): Stress en bonos, vigilar contagio')
        if japan and japan.get('condition') == 'CARRY_UNWIND_ALERT':
            conclusion_lines.append('- ALERTA JAPON: Posible unwind de carry trade, riesgo de contagio global')
        if net_liq.get('change_1m_pct', 0) < -2:
            conclusion_lines.append(f'- Liquidez cayendo ({net_liq.get("change_1m_pct", 0):.1f}% mensual): Fed drenando')

        conclusion_lines.append('')
        conclusion_lines.append('HACIA DONDE VA:')
        if net_liq.get('regime') == 'EXPANDING':
            conclusion_lines.append('- Liquidez en expansion: Entorno mejorando, momentum deberia funcionar')
        elif net_liq.get('regime') == 'CONTRACTING':
            conclusion_lines.append('- Liquidez contrayendose: QT activo, favorecer calidad sobre cantidad')
        else:
            conclusion_lines.append('- Liquidez estable: Mantener enfoque balanceado')

        for line in conclusion_lines:
            if line.startswith('SITUACION') or line.startswith('ESTAR ATENTOS') or line.startswith('HACIA DONDE'):
                ws_mp.cell(row=row, column=1, value=line).font = Font(bold=True)
            elif line.startswith('-'):
                ws_mp.cell(row=row, column=1, value=line).font = Font(size=10)
            else:
                ws_mp.cell(row=row, column=1, value=line)
            row += 1

        row += 1

        # =================================================================
        # SECCION 5: AJUSTES DE PESO APLICADOS
        # =================================================================
        ws_mp.cell(row=row, column=1, value='AJUSTES DE PESO APLICADOS').font = Font(bold=True, size=12, color=DARK_BLUE)
        ws_mp.cell(row=row, column=1).fill = light_blue_fill
        row += 2

        adj = ma.get('weight_adjustments', {})
        if adj:
            for factor, change in adj.items():
                ws_mp.cell(row=row, column=1, value=factor.capitalize())
                cell = ws_mp.cell(row=row, column=2, value=f"{change:+.0%}")
                if change > 0:
                    cell.fill = green_fill
                elif change < 0:
                    cell.fill = red_fill
                row += 1
        else:
            ws_mp.cell(row=row, column=1, value='Sin ajustes (regimen neutral)')
            row += 1

    else:
        ws_mp['A4'] = 'Analisis monetario no disponible'
        ws_mp['A5'] = 'El modulo de fontaneria monetaria no pudo obtener datos.'
        ws_mp['A6'] = 'Se usan pesos por defecto.'

    set_col_widths(ws_mp, [35, 25, 20, 10, 12])

    # =========================================================================
    # HOJA: TRIGGERS ACTIVOS (Mejorado con todos los trades de Congress)
    # =========================================================================
    ws_trig = wb.create_sheet('Triggers_Activos')

    ws_trig['A1'] = 'CONGRESS TRADES - Todos los trades de congresistas'
    ws_trig['A1'].font = title_font
    ws_trig.merge_cells('A1:H1')

    ws_trig['A2'] = 'Historial completo de trades de congresistas. Los tickers donde MAS congresistas coinciden tienen mayor probabilidad de senal.'
    ws_trig['A2'].font = Font(italic=True, size=10)

    # Obtener TODOS los trades de la base de datos
    all_congress_trades = []
    if SIGNAL_DB_AVAILABLE:
        try:
            db = SignalDatabase()
            all_congress_trades = db.get_recent_congress_trades(days=90)
        except Exception as e:
            print(f"  Error obteniendo trades de DB: {e}")

    # Si no hay de DB, usar los de TRIGGERED_STOCKS
    if not all_congress_trades:
        for ticker, triggers_list in TRIGGERED_STOCKS.items():
            for t in triggers_list:
                if t.get('type') == 'CONGRESS_TRADE':
                    all_congress_trades.append({
                        'ticker': ticker,
                        'politician': t.get('politician', 'N/A'),
                        'type': 'PURCHASE' if t.get('action') == 'BUY' else 'SALE',
                        'transaction_date': t.get('date', ''),
                        'amount_range': t.get('amount', ''),
                        'party': t.get('party', ''),
                    })

    # =================================================================
    # SECCION 1: RANKING DE TICKERS MAS POPULARES ENTRE CONGRESISTAS
    # =================================================================
    row = 4
    ws_trig.cell(row=row, column=1, value='TICKERS MAS POPULARES ENTRE CONGRESISTAS').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_trig.cell(row=row, column=1).fill = light_blue_fill
    row += 1
    ws_trig.cell(row=row, column=1, value='Tickers donde multiples congresistas coinciden. Mayor coincidencia = senal mas fuerte.').font = Font(italic=True, size=9)
    row += 2

    # Agrupar por ticker y contar
    ticker_counts = Counter()
    ticker_buys = Counter()
    ticker_sells = Counter()
    ticker_politicians = {}

    for trade in all_congress_trades:
        ticker = trade.get('ticker', '')
        if not ticker or ticker == '--':
            continue
        ticker_counts[ticker] += 1
        action = trade.get('type', trade.get('action', ''))
        if 'PURCHASE' in str(action).upper() or 'BUY' in str(action).upper():
            ticker_buys[ticker] += 1
        elif 'SALE' in str(action).upper() or 'SELL' in str(action).upper():
            ticker_sells[ticker] += 1
        # Track politicians
        pol = trade.get('politician', '')
        if ticker not in ticker_politicians:
            ticker_politicians[ticker] = set()
        if pol:
            ticker_politicians[ticker].add(pol)

    # Headers para ranking
    rank_headers = ['Rank', 'Ticker', 'Total Trades', 'Compras', 'Ventas', 'Ratio C/V', 'Politicos', 'Senal', 'Link']
    for col, h in enumerate(rank_headers, 1):
        cell = ws_trig.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    row += 1

    # Top 30 tickers más populares
    top_tickers = ticker_counts.most_common(30)
    rank = 1
    for ticker, count in top_tickers:
        buys = ticker_buys[ticker]
        sells = ticker_sells[ticker]
        politicians = ticker_politicians.get(ticker, set())

        ws_trig.cell(row=row, column=1, value=rank)
        ws_trig.cell(row=row, column=2, value=ticker).font = Font(bold=True)
        ws_trig.cell(row=row, column=3, value=count)
        ws_trig.cell(row=row, column=4, value=buys)
        ws_trig.cell(row=row, column=5, value=sells)

        # Ratio compras/ventas
        if sells > 0:
            ratio = buys / sells
            ratio_str = f"{ratio:.1f}x"
        else:
            ratio = 999 if buys > 0 else 0
            ratio_str = "Solo compras" if buys > 0 else "-"
        ws_trig.cell(row=row, column=6, value=ratio_str)

        # Número de políticos distintos
        ws_trig.cell(row=row, column=7, value=len(politicians))

        # Señal basada en ratio
        if buys > sells and count >= 3:
            signal = 'BULLISH'
            ws_trig.cell(row=row, column=8, value=signal).fill = green_fill
        elif sells > buys and count >= 3:
            signal = 'BEARISH'
            ws_trig.cell(row=row, column=8, value=signal).fill = red_fill
        else:
            signal = 'NEUTRAL'
            ws_trig.cell(row=row, column=8, value=signal)

        # Link
        link_url = f"https://housestockwatcher.com/summary_by_ticker/{ticker}"
        ws_trig.cell(row=row, column=9, value=link_url).font = Font(size=8, color='0000FF', underline='single')

        for col in range(1, 10):
            ws_trig.cell(row=row, column=col).border = thin_border
        row += 1
        rank += 1

    if not top_tickers:
        ws_trig.cell(row=row, column=1, value='No hay trades de congresistas en la base de datos.')
        ws_trig.cell(row=row, column=1).font = Font(italic=True)
        row += 1

    row += 2

    # =================================================================
    # SECCION 2: DETALLE DE TODOS LOS TRADES RECIENTES
    # =================================================================
    ws_trig.cell(row=row, column=1, value='DETALLE DE TRADES RECIENTES (ultimos 90 dias)').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_trig.cell(row=row, column=1).fill = light_blue_fill
    row += 2

    detail_headers = ['Ticker', 'Politico', 'Partido', 'Accion', 'Monto', 'Fecha', 'Camara', 'Link']
    for col, h in enumerate(detail_headers, 1):
        cell = ws_trig.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_font
    row += 1

    # Ordenar por fecha más reciente
    try:
        all_congress_trades.sort(key=lambda x: x.get('transaction_date', ''), reverse=True)
    except:
        pass

    trades_shown = 0
    for trade in all_congress_trades[:100]:  # Mostrar hasta 100 trades
        ticker = trade.get('ticker', '')
        if not ticker or ticker == '--':
            continue

        ws_trig.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        ws_trig.cell(row=row, column=2, value=trade.get('politician', 'N/A'))

        # Partido con color
        party = trade.get('party', '')
        cell_party = ws_trig.cell(row=row, column=3, value=party)
        if party == 'D':
            cell_party.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        elif party == 'R':
            cell_party.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

        # Acción con color
        action = trade.get('type', trade.get('action', ''))
        if 'PURCHASE' in str(action).upper() or 'BUY' in str(action).upper():
            cell_action = ws_trig.cell(row=row, column=4, value='COMPRA')
            cell_action.fill = green_fill
        elif 'SALE' in str(action).upper() or 'SELL' in str(action).upper():
            cell_action = ws_trig.cell(row=row, column=4, value='VENTA')
            cell_action.fill = red_fill
        else:
            ws_trig.cell(row=row, column=4, value=action)

        ws_trig.cell(row=row, column=5, value=trade.get('amount_range', trade.get('amount', '')))
        ws_trig.cell(row=row, column=6, value=str(trade.get('transaction_date', ''))[:10])
        ws_trig.cell(row=row, column=7, value=trade.get('chamber', ''))

        link_url = f"https://housestockwatcher.com/summary_by_ticker/{ticker}"
        ws_trig.cell(row=row, column=8, value=link_url).font = Font(size=8, color='0000FF', underline='single')

        for col in range(1, 9):
            ws_trig.cell(row=row, column=col).border = thin_border
        row += 1
        trades_shown += 1

    if trades_shown == 0:
        ws_trig.cell(row=row, column=1, value='No hay trades detallados disponibles.')
        ws_trig.cell(row=row, column=1).font = Font(italic=True)
        row += 1

    row += 2

    # =================================================================
    # SECCION 3: CORRELACIONES (mantenemos pero simplificado)
    # =================================================================
    ws_trig.cell(row=row, column=1, value='CORRELACIONES DE MERCADO ACTIVADAS').font = Font(bold=True, size=12, color=DARK_BLUE)
    ws_trig.cell(row=row, column=1).fill = light_blue_fill
    row += 2

    corr_found = False
    seen_sources = set()
    for ticker, triggers_list in TRIGGERED_STOCKS.items():
        for t in triggers_list:
            if t.get('type') == 'CORRELATION':
                source = t.get('source', '')
                if source not in seen_sources:
                    seen_sources.add(source)
                    corr_found = True
                    ws_trig.cell(row=row, column=1, value=source).font = Font(bold=True)
                    ws_trig.cell(row=row, column=2, value=t.get('move', ''))
                    cell_dir = ws_trig.cell(row=row, column=3, value=t.get('direction', ''))
                    if t.get('direction') == 'UP':
                        cell_dir.fill = green_fill
                    else:
                        cell_dir.fill = red_fill
                    ws_trig.cell(row=row, column=4, value=', '.join(t.get('affected_tickers', [])[:8]))
                    for col in range(1, 5):
                        ws_trig.cell(row=row, column=col).border = thin_border
                    row += 1

    if not corr_found:
        ws_trig.cell(row=row, column=1, value='No hay correlaciones activadas actualmente.')
        ws_trig.cell(row=row, column=1).font = Font(italic=True)
        row += 1

    set_col_widths(ws_trig, [6, 10, 20, 8, 10, 18, 12, 10, 45])

    # =========================================================================
    # HOJA: GUIA DE ESTRATEGIAS
    # =========================================================================
    ws_strat = wb.create_sheet('Estrategias_Guia')

    ws_strat['A1'] = 'GUIA DE ESTRATEGIAS - Que hacer segun tu horizonte temporal'
    ws_strat['A1'].font = title_font
    ws_strat.merge_cells('A1:E1')

    ws_strat['A3'] = 'La estrategia optima depende de tu horizonte. NO es lo mismo invertir a 1 semana que a 5 anos.'
    ws_strat['A3'].font = Font(italic=True, bold=True, size=11)

    row = 5
    if TRIGGER_SYSTEM_AVAILABLE:
        for horizon, strategy in STRATEGY_RECOMMENDATIONS.items():
            # Título horizonte
            ws_strat.cell(row=row, column=1, value=f"{strategy['name']} ({strategy['holding_period']})")
            ws_strat.cell(row=row, column=1).font = Font(bold=True, size=12, color=DARK_BLUE)
            ws_strat.cell(row=row, column=1).fill = light_blue_fill
            ws_strat.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

            # Descripción
            ws_strat.cell(row=row, column=1, value=strategy['description'])
            ws_strat.cell(row=row, column=1).font = Font(italic=True)
            ws_strat.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 2

            # Tabla de parámetros
            params = [
                ['Tamano posicion', strategy['position_sizing']],
                ['Stop Loss', strategy['stop_loss']],
                ['Take Profit', strategy['take_profit']],
                ['Rebalanceo', strategy['rebalance']],
                ['Senales a usar', ', '.join(strategy['signals_to_use'])],
                ['Mejor para', strategy['best_for']],
                ['Funciona cuando', strategy['when_works']],
                ['Falla cuando', strategy['when_fails']],
            ]

            for param, value in params:
                ws_strat.cell(row=row, column=1, value=param).font = Font(bold=True)
                ws_strat.cell(row=row, column=2, value=value)
                ws_strat.cell(row=row, column=1).border = thin_border
                ws_strat.cell(row=row, column=2).border = thin_border
                row += 1

            # Pesos recomendados
            row += 1
            ws_strat.cell(row=row, column=1, value='Pesos recomendados:').font = Font(bold=True, underline='single')
            row += 1
            for factor, weight in strategy['weights'].items():
                ws_strat.cell(row=row, column=1, value=factor.capitalize())
                ws_strat.cell(row=row, column=2, value=f"{weight*100:.0f}%")
                row += 1

            row += 2  # Espacio entre estrategias
    else:
        ws_strat.cell(row=row, column=1, value='Sistema de triggers no disponible. Instalar dependencias.')

    # Recomendación final
    row += 2
    ws_strat.cell(row=row, column=1, value='RECOMENDACION FINAL').font = Font(bold=True, size=12)
    ws_strat.cell(row=row, column=1).fill = green_fill
    row += 2
    recommendations = [
        'Para la MAYORIA de inversores: Estrategia LONG-TERM (Value+Quality)',
        '  - Menor estres, mejor rendimiento historico ajustado por riesgo',
        '  - Solo requiere revision trimestral',
        '  - Funciona en cualquier condicion de mercado (a largo plazo)',
        '',
        'Para traders ACTIVOS: Combinar MEDIUM-TERM con triggers de CONGRESS',
        '  - Los trades de congresistas son un "edge" real (informacion privilegiada legal)',
        '  - Seguir correlaciones de mercado para timing',
        '',
        'EVITAR: Trading muy corto plazo a menos que seas profesional',
        '  - Comisiones y spread erosionan ganancias',
        '  - Requiere dedicacion de tiempo completo',
        '  - La mayoria pierde dinero en day trading',
    ]
    for rec in recommendations:
        ws_strat.cell(row=row, column=1, value=rec)
        row += 1

    set_col_widths(ws_strat, [25, 40, 20, 20])

    # =========================================================================
    # HOJA FINAL: GLOSARIO (Ultima pestaña con todos los ratios)
    # =========================================================================
    ws_gloss = wb.create_sheet('Glosario')

    ws_gloss['A1'] = 'GLOSARIO COMPLETO - Metricas, Ratios y Calculos del Sistema'
    ws_gloss['A1'].font = title_font
    ws_gloss.merge_cells('A1:D1')

    ws_gloss['A2'] = 'Referencia rapida de todas las metricas usadas en el analisis. Usa este glosario para entender cada columna y ratio.'
    ws_gloss['A2'].font = Font(italic=True, size=10)

    glossary_data = [
        ['', '', '', ''],
        # =================================================================
        # RATIOS DE VALORACION
        # =================================================================
        ['RATIOS DE VALORACION', '', '', ''],
        ['Ratio', 'Que mide', 'Bueno si...', 'Como se usa en el score'],
        ['P/E (Price/Earnings)', 'Precio / Beneficio por accion', 'Bajo (<15 barato, >25 caro)', 'Percentil: menor P/E = mejor Value score'],
        ['Fwd P/E', 'P/E usando beneficio estimado proximo ano', 'Bajo (<15 barato)', 'Mas relevante que P/E historico'],
        ['P/B (Price/Book)', 'Precio / Valor contable por accion', 'Bajo (<2 barato, >5 caro)', 'Util para bancos y value investing'],
        ['EV/EBITDA', 'Valor empresa / Beneficio operativo', 'Bajo (<10 barato, >15 caro)', 'Mejor que P/E para comparar empresas con distinta deuda'],
        ['EV/Sales', 'Valor empresa / Ventas', 'Bajo (<2 barato, >5 caro)', 'Util para empresas sin beneficios (growth)'],
        ['FCF Yield', 'Free Cash Flow / Market Cap en %', 'Alto (>5% bueno, >8% excelente)', 'Percentil: mayor FCF Yield = mejor Value score'],
        ['PEG Ratio', 'P/E dividido por crecimiento esperado', 'Bajo (<1 infravalorado, >2 caro)', 'Ajusta P/E por crecimiento - util para growth'],
        ['Dividend Yield', 'Dividendo anual / Precio en %', 'Alto (>3% atractivo)', 'Para estrategias de income'],
        ['', '', '', ''],
        # =================================================================
        # RATIOS DE CALIDAD / RENTABILIDAD
        # =================================================================
        ['RATIOS DE CALIDAD Y RENTABILIDAD', '', '', ''],
        ['Ratio', 'Que mide', 'Bueno si...', 'Como se usa en el score'],
        ['ROE (Return on Equity)', 'Beneficio / Patrimonio neto en %', 'Alto (>15% bueno, >20% excelente)', 'Percentil: mayor ROE = mejor Quality score'],
        ['ROIC (Return on Invested Capital)', 'Beneficio / Capital total invertido', 'Alto (>15% excelente)', 'Mejor que ROE - evita distorsion por deuda'],
        ['ROA (Return on Assets)', 'Beneficio / Activos totales', 'Alto (>8% bueno)', 'Mide eficiencia del uso de activos'],
        ['Margen Bruto', '(Ventas - Coste) / Ventas en %', 'Alto (>40% bueno)', 'Poder de pricing'],
        ['Margen Operativo', 'EBIT / Ventas en %', 'Alto (>20% excelente, >10% bueno)', 'Eficiencia operativa - clave para Quality'],
        ['Margen Neto', 'Beneficio Neto / Ventas en %', 'Alto (>10% bueno)', 'Rentabilidad final'],
        ['Asset Turnover', 'Ventas / Activos', 'Alto (>1x eficiente)', 'Eficiencia en uso de activos'],
        ['', '', '', ''],
        # =================================================================
        # RATIOS DE DEUDA / SOLVENCIA
        # =================================================================
        ['RATIOS DE DEUDA Y SOLVENCIA', '', '', ''],
        ['Ratio', 'Que mide', 'Bueno si...', 'Riesgo asociado'],
        ['Deuda/EBITDA', 'Deuda total / EBITDA', 'Bajo (<2x saludable, >4x riesgoso)', 'Percentil: menor = mejor Quality score'],
        ['Net Debt/EBITDA', 'Deuda neta (deuda - caja) / EBITDA', 'Bajo (<2x ideal, <0 = caja neta)', 'Mas preciso que Deuda/EBITDA'],
        ['Debt/Equity', 'Deuda / Patrimonio neto', 'Bajo (<1x conservador, >2x apalancado)', 'Mide apalancamiento financiero'],
        ['Interest Coverage', 'EBIT / Gastos financieros', 'Alto (>5x seguro, <2x peligroso)', 'Capacidad de pagar intereses'],
        ['Quick Ratio', '(Activo corriente - Inventario) / Pasivo corriente', 'Alto (>1x seguro)', 'Liquidez a corto plazo'],
        ['Current Ratio', 'Activo corriente / Pasivo corriente', 'Alto (>1.5x seguro, <1x peligro)', 'Solvencia a corto plazo'],
        ['Cash Ratio', 'Caja / Pasivo corriente', 'Alto (>0.5x muy liquido)', 'Liquidez maxima'],
        ['', '', '', ''],
        # =================================================================
        # RATIOS DE LIQUIDEZ DEL MERCADO
        # =================================================================
        ['RATIOS DE LIQUIDEZ DE MERCADO (Monetary Plumbing)', '', '', ''],
        ['Indicador', 'Que mide', 'Bueno para RV si...', 'Como afecta al scoring'],
        ['Net Liquidity', 'Fed Balance - TGA - RRP en billones', 'Subiendo (inyeccion)', '+peso a Momentum y Growth'],
        ['VIX', 'Volatilidad implicita del S&P500', 'Bajo (<15 calma, >25 miedo)', 'VIX alto = -peso a Momentum'],
        ['MOVE Index', 'VIX de los bonos del Tesoro', 'Bajo (<80 calma, >120 stress)', 'MOVE alto = +peso a Quality'],
        ['Credit Spreads', 'Diferencia HYG-LQD en bps', 'Bajo (<300 normal, >500 crisis)', 'Spreads altos = modo defensivo'],
        ['USD/JPY', 'Dolares por Yen japones', 'Estable (carry trade tranquilo)', 'Caida rapida = riesgo de crash'],
        ['TGA (Treasury General Account)', 'Cuenta del Tesoro en la Fed', 'Bajando (gasto fiscal)', 'TGA baja = liquidez sube'],
        ['RRP (Reverse Repo)', 'Dinero aparcado en la Fed', 'Bajando (sale al mercado)', 'RRP baja = liquidez disponible'],
        ['', '', '', ''],
        # =================================================================
        # METRICAS DE MOMENTUM
        # =================================================================
        ['METRICAS DE MOMENTUM', '', '', ''],
        ['Metrica', 'Que mide', 'Bueno si...', 'Como se usa'],
        ['Mom 12M', 'Retorno ultimos 12 meses %', 'Positivo y fuerte (>20%)', 'Percentil: mayor retorno = mejor Momentum'],
        ['Mom 6M', 'Retorno ultimos 6 meses %', 'Positivo', 'Confirma tendencia'],
        ['Mom 1M', 'Retorno ultimo mes %', 'Positivo', 'Momentum corto plazo'],
        ['RSI 14', 'Relative Strength Index (14 dias)', 'Entre 30-70 (>70 sobrecompra, <30 sobreventa)', 'Para timing de entrada'],
        ['MACD', 'Diferencia entre EMAs 12 y 26', 'Cruce alcista', 'Confirma cambio de tendencia'],
        ['Analyst Revisions', 'Cambio en estimaciones de analistas', 'Revisiones al alza', 'Percentil: mas revisiones = mejor score'],
        ['Price vs SMA 200', 'Precio relativo a media movil 200d', 'Por encima (tendencia alcista)', 'Confirma tendencia largo plazo'],
        ['', '', '', ''],
        # =================================================================
        # METRICAS DE RIESGO
        # =================================================================
        ['METRICAS DE RIESGO (LowVol Score)', '', '', ''],
        ['Metrica', 'Que mide', 'Bueno si...', 'Como se usa'],
        ['Beta', 'Sensibilidad al mercado vs S&P500', 'Bajo (<1 defensivo, >1.5 agresivo)', 'Percentil: menor Beta = mejor LowVol'],
        ['Volatilidad 252d', 'Desviacion estandar anualizada %', 'Bajo (<25% estable, >40% volatil)', 'Percentil: menor vol = mejor LowVol'],
        ['Max Drawdown', 'Maxima caida desde maximo', 'Bajo (<-20% resiliente)', 'Riesgo de perdida maxima'],
        ['Sharpe Ratio', 'Retorno / Volatilidad', 'Alto (>1 bueno, >2 excelente)', 'Retorno ajustado por riesgo'],
        ['', '', '', ''],
        # =================================================================
        # FACTORES ESPECULATIVOS
        # =================================================================
        ['FACTORES ESPECULATIVOS', '', '', ''],
        ['Factor', 'Que detecta', 'Bueno si...', 'Como se usa'],
        ['Congress Score', 'Trades de congresistas USA', 'Compras recientes >100k USD', '+puntos si compran, -si venden'],
        ['Polymarket Score', 'Smart money en mercados prediccion', 'Apuestas grandes a favor', '+puntos si hay senales positivas'],
        ['Insider Buying', 'Compras de directivos', 'Compras significativas', 'Insiders saben mas que nosotros'],
        ['Short Interest', '% de acciones en corto', 'Bajo (<5% normal, >20% short squeeze)', 'Alto puede ser oportunidad o riesgo'],
        ['', '', '', ''],
        # =================================================================
        # SENALES DEL SISTEMA
        # =================================================================
        ['SENALES FINALES DEL SISTEMA', '', '', ''],
        ['Senal', 'Criterio', 'Accion recomendada', ''],
        ['STRONG BUY', 'Score >= 60 Y Upside >= 10%', 'Comprar agresivamente, posicion completa', ''],
        ['BUY', 'Score >= 60 Y Upside >= 0%', 'Comprar, empezar posicion', ''],
        ['ACCUMULATE', 'Score >= 50 Y Upside >= 0%', 'Ir acumulando poco a poco', ''],
        ['HOLD', 'Score 40-60 O Upside negativo', 'Mantener si tienes, no anadir', ''],
        ['REDUCE', 'Score 30-40', 'Reducir posicion gradualmente', ''],
        ['SELL', 'Score < 30', 'Vender posicion', ''],
        ['', '', '', ''],
        # =================================================================
        # CALCULO DEL SCORE
        # =================================================================
        ['CALCULO DEL SCORE COMPUESTO', '', '', ''],
        ['Componente', 'Descripcion', 'Peso Default', 'Ajuste por Liquidez'],
        ['Value Score', 'Percentil de valoracion (P/E, EV/EBITDA, FCF)', '20%', '+5% en liquidez baja'],
        ['Quality Score', 'Percentil de calidad (ROE, Margen, Deuda)', '25%', '+10% en crisis'],
        ['Momentum Score', 'Percentil de momentum (12M return, revisions)', '15%', '+10% en liquidez alta'],
        ['LowVol Score', 'Percentil de baja volatilidad (Beta, Vol)', '15%', '+5% en crisis'],
        ['Congress Score', 'Ajuste por trades de congresistas', '10%', 'Sin ajuste'],
        ['Polymarket Score', 'Ajuste por smart money', '10%', 'Sin ajuste'],
        ['', '', '', ''],
        ['FORMULA FINAL:', 'Score = Sum(Factor_i * Peso_i) ajustado por regimen macro', '', ''],
        ['IMPORTANTE:', 'El Upside (% al target) es CRITICO - Upside negativo = NUNCA es BUY', '', ''],
        ['', '', '', ''],
        # =================================================================
        # REFERENCIAS UTILES
        # =================================================================
        ['REFERENCIAS Y FUENTES', '', '', ''],
        ['Fuente', 'URL', 'Datos que provee', ''],
        ['Yahoo Finance', 'finance.yahoo.com', 'Precios, fundamentales, targets', ''],
        ['House Stock Watcher', 'housestockwatcher.com', 'Trades de representantes USA', ''],
        ['Senate Stock Watcher', 'senatestockwatcher.com', 'Trades de senadores USA', ''],
        ['FRED (Federal Reserve)', 'fred.stlouisfed.org', 'Balance Fed, TGA, tipos', ''],
        ['Polymarket', 'polymarket.com', 'Mercados de prediccion', ''],
        ['Finnhub', 'finnhub.io', 'Noticias, sentimiento', ''],
    ]

    for i, row_data in enumerate(glossary_data):
        for col, val in enumerate(row_data, 1):
            cell = ws_gloss.cell(row=i+4, column=col, value=val)
            # Seccion headers (en mayusculas)
            if val and val.isupper() and len(val) > 10 and 'RATIOS' in val or 'METRICAS' in val or 'FACTORES' in val or 'SENALES' in val or 'CALCULO' in val or 'REFERENCIAS' in val:
                cell.font = Font(bold=True, size=12, color=DARK_BLUE)
                cell.fill = light_blue_fill
            # Column headers
            elif val in ['Ratio', 'Metrica', 'Factor', 'Senal', 'Componente', 'Indicador', 'Fuente',
                        'Que mide', 'Bueno si...', 'Como se usa', 'Como se usa en el score',
                        'Riesgo asociado', 'Bueno para RV si...', 'Como afecta al scoring',
                        'Accion recomendada', 'Criterio', 'Descripcion', 'Peso Default',
                        'Ajuste por Liquidez', 'URL', 'Datos que provee']:
                cell.fill = header_fill
                cell.font = white_font
            # Important notes
            elif val and ('IMPORTANTE' in val or 'FORMULA' in val):
                cell.font = Font(bold=True, color='9C0006')
            # Links
            elif val and ('finance.yahoo' in val or 'stockwatcher' in val or 'fred.' in val or 'polymarket' in val or 'finnhub' in val):
                cell.font = Font(color='0000FF', underline='single')
            cell.border = thin_border

    set_col_widths(ws_gloss, [25, 45, 35, 30])

    # Guardar
    wb.save(output_path)
    return wb


# =============================================================================
# MAIN
# =============================================================================

# Variable global para almacenar info de triggers
TRIGGERED_STOCKS = {}
TRIGGER_SUMMARY = []


def main():
    global TRIGGERED_STOCKS, TRIGGER_SUMMARY

    parser = argparse.ArgumentParser(description='Market Analyzer - Sistema Semi-Automatizado')
    parser.add_argument('--regime', choices=['RECOVERY', 'EXPANSION', 'LATE_EXPANSION', 'CONTRACTION'],
                        help='Cambiar régimen macro')
    parser.add_argument('--risk', choices=['RISK_ON', 'NEUTRAL', 'RISK_OFF'],
                        help='Cambiar risk appetite')
    parser.add_argument('--output', default='market_analysis_auto.xlsx',
                        help='Nombre del archivo de salida')
    parser.add_argument('--add', nargs='+', help='Añadir tickers al sector Otros')
    parser.add_argument('--no-triggers', action='store_true',
                        help='Desactivar escaneo de triggers')

    args = parser.parse_args()

    # Actualizar configuración si se especifica
    if args.regime:
        MACRO_CONFIG['regime'] = args.regime
    if args.risk:
        MACRO_CONFIG['risk_appetite'] = args.risk
    if args.add:
        if 'Otros' not in UNIVERSE:
            UNIVERSE['Otros'] = []
        UNIVERSE['Otros'].extend(args.add)

    print("=" * 60)
    print("MARKET ANALYZER V14 - Con Sistema de Triggers")
    print("=" * 60)
    print(f"Regimen: {MACRO_CONFIG['regime']}")
    print(f"Risk Appetite: {MACRO_CONFIG['risk_appetite']}")
    print(f"Pesos: {get_final_weights()}")

    # PASO 1: Escanear triggers para descubrir nuevas oportunidades
    if TRIGGER_SYSTEM_AVAILABLE and not args.no_triggers:
        print("\n" + "=" * 60)
        print("FASE 1: ESCANEO DE TRIGGERS")
        print("=" * 60)
        TRIGGERED_STOCKS, TRIGGER_SUMMARY = run_full_scan()

        # Añadir stocks triggered al universo (si no están ya)
        existing_tickers = set()
        for sector_tickers in UNIVERSE.values():
            existing_tickers.update(sector_tickers)

        new_triggered = []
        for ticker in TRIGGERED_STOCKS.keys():
            if ticker not in existing_tickers and len(ticker) <= 5:
                new_triggered.append(ticker)

        if new_triggered:
            print(f"\n  Añadiendo {len(new_triggered)} stocks triggered al análisis...")
            if 'Triggered' not in UNIVERSE:
                UNIVERSE['Triggered'] = []
            UNIVERSE['Triggered'].extend(new_triggered[:20])  # Máx 20 nuevos

    print("\n" + "=" * 60)
    print("FASE 2: OBTENCION DE DATOS")
    print("=" * 60)

    # Obtener datos
    companies = fetch_all_data()
    
    if not companies:
        print("\n❌ No se pudieron obtener datos. Verifica conexión a internet.")
        sys.exit(1)
    
    # Calcular scores
    print("\n" + "=" * 60)
    print("FASE 3: CALCULO DE SCORES")
    print("=" * 60)
    companies = calculate_scores(companies)

    # Almacenar en base de datos local (para histórico y backtesting futuro)
    if SIGNAL_DB_AVAILABLE:
        try:
            db = SignalDatabase()
            today = datetime.now().strftime('%Y-%m-%d')
            stored = db.store_scores_bulk(companies, today)
            print(f"\n  Base de datos: {stored} scores almacenados para {today}")
            stats = db.get_database_stats()
            print(f"  Total en DB: {stats.get('calculated_scores', 0)} registros de scores")
        except Exception as e:
            print(f"  Warning: No se pudo almacenar en DB: {e}")

    # Generar Excel
    print(f"\n  Generando Excel: {args.output}")
    create_excel(companies, args.output)
    
    # Resumen
    signals = {'STRONG BUY': 0, 'BUY': 0, 'ACCUMULATE': 0, 'HOLD': 0, 'REDUCE': 0, 'SELL': 0}
    for c in companies:
        if c['signal'] in signals:
            signals[c['signal']] += 1
    
    print("\n" + "=" * 60)
    print("✅ COMPLETADO")
    print("=" * 60)
    print(f"\n📋 Distribucion de senales:")
    for sig, count in signals.items():
        emoji = {'STRONG BUY': '🟢🟢', 'BUY': '🟢', 'ACCUMULATE': '🟡', 'HOLD': '⚪', 'REDUCE': '🟠', 'SELL': '🔴'}
        print(f"   {emoji.get(sig, '')} {sig}: {count}")
    
    print(f"\n🎯 TOP 5 PICKS:")
    for c in companies[:5]:
        print(f"   {c['ticker']:8} | Score: {c['composite_score']:.1f} | {c['signal']:10} | ${c['price']}")
    
    print(f"\n📁 Archivo guardado: {args.output}")
    print("\n💡 Próximo paso: Abre el Excel y revisa la hoja 'Action_List'")


if __name__ == '__main__':
    main()
